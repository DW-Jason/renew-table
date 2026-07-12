# -*- coding: utf-8 -*-
import os
import sys
import datetime
import json
import math
import time
import threading
import ctypes
import unicodedata
from ctypes import wintypes
from copy import copy
from collections import defaultdict
import tkinter as tk
import tkinter.font as tkfont
from tkinter import messagebox, ttk
import tkinterdnd2
import re

APP_CONFIG_DIR_NAME = "查房表更新工具"
CONFIG_FILENAME = "config.ini"
DEFAULT_WARD = "01"
DEFAULT_GROUP = "默认医疗组"


def get_app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


def get_user_config_dir():
    base_dir = (
        os.environ.get("LOCALAPPDATA")
        or os.environ.get("APPDATA")
        or os.path.expanduser("~")
    )
    return os.path.join(base_dir, APP_CONFIG_DIR_NAME)


def get_config_file():
    if getattr(sys, "frozen", False):
        return os.path.join(get_user_config_dir(), CONFIG_FILENAME)
    return os.path.join(get_app_dir(), CONFIG_FILENAME)


def get_legacy_config_files():
    if not getattr(sys, "frozen", False):
        return []
    return [
        os.path.join(get_app_dir(), CONFIG_FILENAME),
        os.path.join(os.getcwd(), CONFIG_FILENAME),
    ]


CONFIG_FILE = get_config_file()
EXPORT_A_RECORDING_FILE = os.path.join(
    os.path.dirname(CONFIG_FILE),
    "export_a_recording.json",
)
APP_WINDOW_TITLE = "查房表更新工具"
EXPORT_A_WINDOW_KEYWORDS = [
    "医生工作站",
    "医师工作站",
    "病员管理",
    "福建省立医院",
    "胃肠外科",
]


def get_resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        return os.path.join(sys._MEIPASS, relative_path)
    return relative_path

def load_config():
    cfg = {
        'row_height': 30,
        'ward': DEFAULT_WARD,
        'beds': [],
        'groups': {DEFAULT_GROUP: []},
        'current_group': DEFAULT_GROUP,
    }
    config_file = CONFIG_FILE
    if not os.path.exists(config_file):
        for legacy_file in get_legacy_config_files():
            if os.path.exists(legacy_file):
                config_file = legacy_file
                break

    if os.path.exists(config_file):
        try:
            import configparser
            cp = configparser.ConfigParser(interpolation=None)
            cp.read(config_file, encoding='utf-8')
            cfg['row_height'] = cp.getint('Settings', 'row_height', fallback=30)
            ward = cp.get('Settings', 'ward', fallback=DEFAULT_WARD).strip()
            if ward.isdigit():
                ward_int = int(ward)
                if 1 <= ward_int <= 99:
                    cfg['ward'] = f"{ward_int:02d}"
            beds_text = cp.get('Settings', 'beds', fallback='')
            beds = []
            for item in beds_text.split(','):
                item = item.strip()
                if item.isdigit():
                    bed_int = int(item)
                    if 1 <= bed_int <= 50:
                        bed = f"{bed_int:02d}"
                        if bed not in beds:
                            beds.append(bed)
            cfg['beds'] = beds
            groups = {}
            groups_text = cp.get('Settings', 'medical_groups', fallback='').strip()
            if groups_text:
                try:
                    raw_groups = json.loads(groups_text)
                    if isinstance(raw_groups, dict):
                        for group_name, group_beds in raw_groups.items():
                            name = str(group_name).strip()
                            if not name or not isinstance(group_beds, list):
                                continue
                            normalized_beds = []
                            for item in group_beds:
                                item = str(item).strip()
                                if item.isdigit():
                                    bed_int = int(item)
                                    if 1 <= bed_int <= 50:
                                        bed = f"{bed_int:02d}"
                                        if bed not in normalized_beds:
                                            normalized_beds.append(bed)
                            groups[name] = normalized_beds
                except Exception:
                    groups = {}
            if not groups:
                groups = {DEFAULT_GROUP: beds}
            current_group = cp.get('Settings', 'current_group', fallback='').strip()
            if current_group not in groups:
                current_group = next(iter(groups), DEFAULT_GROUP)
            cfg['groups'] = groups
            cfg['current_group'] = current_group
            cfg['beds'] = groups.get(current_group, [])
        except Exception:
            pass
    return cfg

def save_config(row_height=None, ward=None, beds=None, groups=None, current_group=None):
    try:
        import configparser
        cp = configparser.ConfigParser(interpolation=None)
        cp.read(CONFIG_FILE, encoding='utf-8') if os.path.exists(CONFIG_FILE) else None
        if not cp.has_section('Settings'):
            cp.add_section('Settings')
        if row_height is not None:
            cp.set('Settings', 'row_height', str(row_height))
        if ward is not None:
            cp.set('Settings', 'ward', str(ward))
        if beds is not None:
            cp.set('Settings', 'beds', ','.join(beds))
        if groups is not None:
            cp.set('Settings', 'medical_groups', json.dumps(groups, ensure_ascii=False))
        if current_group is not None:
            cp.set('Settings', 'current_group', str(current_group))
        config_dir = os.path.dirname(CONFIG_FILE)
        if config_dir:
            os.makedirs(config_dir, exist_ok=True)
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            cp.write(f)
    except Exception:
        pass


class DropZone(tk.Frame):
    """拖放区：整框任意位置可点、可拖放；事件绑定到自身及全部子控件。"""

    _ui_font_hint = ("Microsoft YaHei UI", "微软雅黑", "Segoe UI")
    _ui_font_path = ("Microsoft YaHei UI", "微软雅黑", "Segoe UI")

    def __init__(self, parent, label_text, color="#FFFFFF", zone_letter=None, **kwargs):
        super().__init__(parent, **kwargs)
        self.label_text = label_text
        self._zone_color = color
        self.file_path = None
        self.zone_letter = zone_letter
        self.parent_app = None
        self._select_after_id = None

        self.configure(
            bd=1,
            relief=tk.SOLID,
            bg="#9DB6CE",
            width=374,
            height=150,
            cursor="hand2",
        )

        self.inner_frame = tk.Frame(self, bg=color, cursor="hand2")
        self.inner_frame.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)

        self.content_frame = tk.Frame(self.inner_frame, bg=color, cursor="hand2")
        self.content_frame.pack(fill=tk.BOTH, expand=True, padx=16, pady=14)

        badge_color = "#1D6F93" if zone_letter == "A" else "#1E7D54"
        self.header_frame = tk.Frame(self.content_frame, bg=color, cursor="hand2")
        self.header_frame.pack(fill=tk.X, pady=(0, 10))

        self.badge_label = tk.Label(
            self.header_frame,
            text=zone_letter or "",
            font=(DropZone._ui_font_hint[0], 15, "bold"),
            bg=badge_color,
            fg="#FFFFFF",
            width=3,
            height=1,
            cursor="hand2",
        )
        self.badge_label.pack(side=tk.LEFT)

        self.hint_label = tk.Label(
            self.content_frame,
            text=self.label_text,
            font=(DropZone._ui_font_hint[0], 12),
            bg=color,
            fg="#07182F",
            justify=tk.LEFT,
            anchor=tk.W,
            cursor="hand2",
        )
        self.hint_label.pack(fill=tk.X)

        self.path_label = tk.Label(
            self.content_frame,
            text="",
            font=(DropZone._ui_font_path[0], 11, "bold"),
            bg=color,
            fg="#246BFE",
            wraplength=320,
            justify=tk.LEFT,
            anchor=tk.W,
            cursor="hand2",
        )
        self.path_label.pack(fill=tk.X, pady=(6, 0))

        self.status_hint_label = tk.Label(
            self.content_frame,
            text="",
            font=(DropZone._ui_font_hint[0], 9, "bold"),
            bg=color,
            fg="#17835B",
            justify=tk.LEFT,
            anchor=tk.W,
            cursor="hand2",
        )
        self.status_hint_label.pack(fill=tk.X, pady=(3, 0))

        self.pack_propagate(False)

        self._apply_font_fallbacks()
        self._bind_tree("<Button-3>", self.show_context_menu)
        self._set_cursor_tree("hand2")

    def _apply_font_fallbacks(self):
        families_avail = tkfont.families()
        for lbl, families, size_tuple in (
            (self.hint_label, DropZone._ui_font_hint, (12,)),
            (self.path_label, DropZone._ui_font_path, (11, "bold")),
        ):
            for fam in families:
                if fam in families_avail:
                    lbl.configure(font=(fam,) + size_tuple)
                    break

    @staticmethod
    def _iter_tree(root_widget):
        yield root_widget
        for ch in root_widget.winfo_children():
            yield from DropZone._iter_tree(ch)

    def _bind_tree(self, sequence, handler):
        for w in self._iter_tree(self):
            w.bind(sequence, handler)

    def _set_cursor_tree(self, cursor_name):
        for w in self._iter_tree(self):
            try:
                w.configure(cursor=cursor_name)
            except tk.TclError:
                pass

    def _cancel_pending_select(self):
        if self._select_after_id is not None:
            self.after_cancel(self._select_after_id)
            self._select_after_id = None

    def bind_select(self, callback):
        """整块区域可点；单击延迟选文件，避免双击打开文件夹时误弹对话框。"""

        def on_single(_e):
            self._cancel_pending_select()

            def fire():
                self._select_after_id = None
                try:
                    callback()
                except TypeError:
                    # 兼容误写成 lambda e: ... 的旧调用方式
                    callback(None)

            self._select_after_id = self.after(280, fire)

        def on_double(_e):
            self._cancel_pending_select()
            self.open_folder()

        for w in self._iter_tree(self):
            w.bind("<Button-1>", on_single)
            w.bind("<Double-Button-1>", on_double)

    def register_dnd(self, on_drop):
        for w in self._iter_tree(self):
            w.drop_target_register(tkinterdnd2.DND_FILES)
            w.dnd_bind("<<Drop>>", on_drop)

    def show_context_menu(self, event):
        if not self.file_path:
            return
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="打开文件位置", command=self.open_folder)
        menu.post(event.x_root, event.y_root)
    
    def open_folder(self):
        if self.file_path and self.parent_app:
            folder = os.path.dirname(self.file_path)
            if os.path.isdir(folder):
                os.startfile(folder)

    def set_file(self, path):
        self.file_path = path
        filename = os.path.basename(path)
        self.path_label.config(text=filename)
        done_border = "#6FB88D"
        done_bg = "#FFFFFF"
        filename_bg = "#EAF8EF"
        self.configure(bg=done_border, relief=tk.SOLID)
        self.inner_frame.configure(bg=done_bg)
        self.content_frame.configure(bg=done_bg)
        self.hint_label.configure(bg=done_bg)
        self.path_label.configure(bg=filename_bg, fg="#146C43", padx=8, pady=3)
        self.status_hint_label.configure(text="已导入，双击可打开所在文件夹", bg=done_bg, fg="#17835B")

    def clear(self):
        self.file_path = None
        self.path_label.config(text="")
        self.status_hint_label.config(text="")
        c = self._zone_color
        self.configure(bg="#9DB6CE", relief=tk.SOLID)
        self.inner_frame.configure(bg=c)
        self.content_frame.configure(bg=c)
        self.hint_label.configure(bg=c)
        self.path_label.configure(bg=c, fg="#246BFE", padx=0, pady=0)
        self.status_hint_label.configure(bg=c)


class WinHookPoint(ctypes.Structure):
    _fields_ = [("x", wintypes.LONG), ("y", wintypes.LONG)]


class WinMouseHookStruct(ctypes.Structure):
    _fields_ = [
        ("pt", WinHookPoint),
        ("mouseData", wintypes.DWORD),
        ("flags", wintypes.DWORD),
        ("time", wintypes.DWORD),
        ("dwExtraInfo", ctypes.c_void_p),
    ]


class WinKeyboardHookStruct(ctypes.Structure):
    _fields_ = [
        ("vkCode", wintypes.DWORD),
        ("scanCode", wintypes.DWORD),
        ("flags", wintypes.DWORD),
        ("time", wintypes.DWORD),
        ("dwExtraInfo", ctypes.c_void_p),
    ]


class WindowsWindowHelper:
    SYSTEM_CLASSES = {
        "Shell_TrayWnd",
        "Shell_SecondaryTrayWnd",
        "Button",
        "Progman",
        "WorkerW",
        "DV2ControlHost",
    }
    SW_RESTORE = 9
    VK_MENU = 0x12
    KEYEVENTF_KEYUP = 0x0002

    @staticmethod
    def _hwnd_value(hwnd):
        value = getattr(hwnd, "value", hwnd)
        try:
            return int(value or 0)
        except Exception:
            return 0

    @classmethod
    def _setup(cls):
        user32 = ctypes.windll.user32
        user32.GetWindowTextLengthW.argtypes = [wintypes.HWND]
        user32.GetWindowTextLengthW.restype = ctypes.c_int
        user32.GetWindowTextW.argtypes = [wintypes.HWND, wintypes.LPWSTR, ctypes.c_int]
        user32.GetWindowTextW.restype = ctypes.c_int
        user32.GetClassNameW.argtypes = [wintypes.HWND, wintypes.LPWSTR, ctypes.c_int]
        user32.GetClassNameW.restype = ctypes.c_int
        user32.GetWindowRect.argtypes = [wintypes.HWND, ctypes.POINTER(wintypes.RECT)]
        user32.GetWindowRect.restype = wintypes.BOOL
        user32.IsWindowVisible.argtypes = [wintypes.HWND]
        user32.IsWindowVisible.restype = wintypes.BOOL
        user32.IsIconic.argtypes = [wintypes.HWND]
        user32.IsIconic.restype = wintypes.BOOL
        user32.ShowWindow.argtypes = [wintypes.HWND, ctypes.c_int]
        user32.SetForegroundWindow.argtypes = [wintypes.HWND]
        user32.SetForegroundWindow.restype = wintypes.BOOL
        user32.BringWindowToTop.argtypes = [wintypes.HWND]
        user32.BringWindowToTop.restype = wintypes.BOOL
        user32.GetForegroundWindow.restype = wintypes.HWND
        return user32

    @classmethod
    def get_window_text(cls, hwnd):
        user32 = cls._setup()
        length = user32.GetWindowTextLengthW(hwnd)
        if length <= 0:
            return ""
        buffer = ctypes.create_unicode_buffer(length + 1)
        user32.GetWindowTextW(hwnd, buffer, length + 1)
        return buffer.value.strip()

    @classmethod
    def get_class_name(cls, hwnd):
        user32 = cls._setup()
        buffer = ctypes.create_unicode_buffer(256)
        user32.GetClassNameW(hwnd, buffer, 256)
        return buffer.value.strip()

    @classmethod
    def get_window_rect(cls, hwnd):
        user32 = cls._setup()
        rect = wintypes.RECT()
        if not user32.GetWindowRect(hwnd, ctypes.byref(rect)):
            return None
        return [int(rect.left), int(rect.top), int(rect.right), int(rect.bottom)]

    @classmethod
    def get_window_info(cls, hwnd, action_index=None):
        if not hwnd:
            return None
        title = cls.get_window_text(hwnd)
        class_name = cls.get_class_name(hwnd)
        rect = cls.get_window_rect(hwnd)
        info = {
            "title": title,
            "class": class_name,
            "rect": rect,
        }
        if action_index is not None:
            info["action_index"] = int(action_index)
        return info

    @classmethod
    def get_foreground_info(cls, action_index=None):
        user32 = cls._setup()
        hwnd = user32.GetForegroundWindow()
        return cls.get_window_info(hwnd, action_index=action_index)

    @classmethod
    def is_system_window(cls, info):
        if not info:
            return True
        title = (info.get("title") or "").strip()
        class_name = (info.get("class") or "").strip()
        if not title:
            return True
        if APP_WINDOW_TITLE in title:
            return True
        return class_name in cls.SYSTEM_CLASSES

    @classmethod
    def _window_matches_keywords(cls, info):
        title = info.get("title") or ""
        return any(keyword in title for keyword in EXPORT_A_WINDOW_KEYWORDS)

    @staticmethod
    def _same_window_info(left, right):
        return (
            (left.get("title") or "") == (right.get("title") or "")
            and (left.get("class") or "") == (right.get("class") or "")
        )

    @classmethod
    def append_unique_window(cls, windows, info):
        if cls.is_system_window(info):
            return
        clean = {
            "title": info.get("title") or "",
            "class": info.get("class") or "",
            "rect": info.get("rect"),
        }
        if "action_index" in info:
            clean["action_index"] = int(info.get("action_index") or 0)
        if not any(cls._same_window_info(existing, clean) for existing in windows):
            windows.append(clean)

    @classmethod
    def get_taskbar_rects(cls):
        user32 = cls._setup()
        rects = []
        enum_proc_type = getattr(ctypes, "WINFUNCTYPE", ctypes.CFUNCTYPE)(
            wintypes.BOOL,
            wintypes.HWND,
            wintypes.LPARAM,
        )

        def enum_proc(hwnd, _lparam):
            class_name = cls.get_class_name(hwnd)
            if class_name in ("Shell_TrayWnd", "Shell_SecondaryTrayWnd"):
                rect = cls.get_window_rect(hwnd)
                if rect:
                    rects.append(rect)
            return True

        user32.EnumWindows(enum_proc_type(enum_proc), 0)
        return rects

    @staticmethod
    def point_in_rect(x, y, rect, padding=2):
        if not rect or len(rect) != 4:
            return False
        left, top, right, bottom = rect
        return left - padding <= x <= right + padding and top - padding <= y <= bottom + padding

    @classmethod
    def action_in_rects(cls, action, rects):
        if action.get("type") != "mouse":
            return False
        try:
            x = int(float(action.get("x", 0)))
            y = int(float(action.get("y", 0)))
        except Exception:
            return False
        return any(cls.point_in_rect(x, y, rect) for rect in rects or [])

    @classmethod
    def select_target_windows(cls, history, actions, taskbar_rects):
        selected = []
        for info in history:
            if cls._window_matches_keywords(info):
                cls.append_unique_window(selected, info)

        taskbar_indexes = [
            idx
            for idx, action in enumerate(actions)
            if action.get("event") == "up" and cls.action_in_rects(action, taskbar_rects)
        ]
        for action_index in taskbar_indexes:
            for info in history:
                if int(info.get("action_index", 0)) >= action_index:
                    cls.append_unique_window(selected, info)
                    break

        for info in history:
            cls.append_unique_window(selected, info)
            if len(selected) >= 8:
                break
        return selected[:8]

    @classmethod
    def enum_visible_windows(cls):
        user32 = cls._setup()
        enum_proc_type = getattr(ctypes, "WINFUNCTYPE", ctypes.CFUNCTYPE)(
            wintypes.BOOL,
            wintypes.HWND,
            wintypes.LPARAM,
        )
        windows = []

        def enum_proc(hwnd, _lparam):
            if user32.IsWindowVisible(hwnd):
                info = cls.get_window_info(hwnd)
                if not cls.is_system_window(info):
                    info["_hwnd"] = hwnd
                    windows.append(info)
            return True

        user32.EnumWindows(enum_proc_type(enum_proc), 0)
        return windows

    @classmethod
    def find_recording_target(cls, recording):
        windows = cls.enum_visible_windows()
        candidates = recording.get("target_windows") or []

        for candidate in candidates:
            title = candidate.get("title") or ""
            class_name = candidate.get("class") or ""
            for window in windows:
                if title and title == (window.get("title") or ""):
                    if not class_name or class_name == (window.get("class") or ""):
                        return candidate, window

        for candidate in candidates:
            title = candidate.get("title") or ""
            if len(title) < 4:
                continue
            for window in windows:
                window_title = window.get("title") or ""
                if title in window_title or window_title in title:
                    return candidate, window

        for keyword in EXPORT_A_WINDOW_KEYWORDS:
            for window in windows:
                if keyword in (window.get("title") or ""):
                    return {}, window

        return None, None

    @classmethod
    def activate_recording_target(cls, recording):
        user32 = cls._setup()
        recorded, current = cls.find_recording_target(recording)
        if not current:
            return None

        hwnd = current.get("_hwnd")
        try:
            user32.ShowWindow(hwnd, cls.SW_RESTORE)
            user32.BringWindowToTop(hwnd)
            user32.keybd_event(cls.VK_MENU, 0, 0, 0)
            user32.SetForegroundWindow(hwnd)
            user32.keybd_event(cls.VK_MENU, 0, cls.KEYEVENTF_KEYUP, 0)
            time.sleep(0.35)
        except Exception:
            pass

        return {
            "recorded": recorded or {},
            "current": {
                "title": current.get("title") or "",
                "class": current.get("class") or "",
                "rect": cls.get_window_rect(hwnd) or current.get("rect"),
            },
        }


class WindowsActionRecorder:
    WH_MOUSE_LL = 14
    WH_KEYBOARD_LL = 13
    HC_ACTION = 0
    PM_REMOVE = 0x0001

    WM_MOUSEMOVE = 0x0200
    WM_LBUTTONDOWN = 0x0201
    WM_LBUTTONUP = 0x0202
    WM_RBUTTONDOWN = 0x0204
    WM_RBUTTONUP = 0x0205
    WM_MBUTTONDOWN = 0x0207
    WM_MBUTTONUP = 0x0208
    WM_MOUSEWHEEL = 0x020A
    WM_KEYDOWN = 0x0100
    WM_KEYUP = 0x0101
    WM_SYSKEYDOWN = 0x0104
    WM_SYSKEYUP = 0x0105
    VK_F8 = 0x77

    MSLLHOOKSTRUCT = WinMouseHookStruct
    KBDLLHOOKSTRUCT = WinKeyboardHookStruct

    def __init__(self, stop_event=None):
        if sys.platform != "win32":
            raise RuntimeError("自动录制仅支持 Windows。")
        self.stop_event = stop_event or threading.Event()
        self.actions = []
        self._last_event_time = None
        self._mouse_hook = None
        self._keyboard_hook = None
        self._pressed_buttons = set()
        self._window_history = []
        self._last_window_key = None
        self._last_window_sample = 0.0
        self._taskbar_rects = []
        self._hook_proc_type = getattr(ctypes, "WINFUNCTYPE", ctypes.CFUNCTYPE)(
            ctypes.c_ssize_t,
            ctypes.c_int,
            wintypes.WPARAM,
            wintypes.LPARAM,
        )
        self._mouse_proc_ref = self._hook_proc_type(self._mouse_proc)
        self._keyboard_proc_ref = self._hook_proc_type(self._keyboard_proc)

    @staticmethod
    def _prepare_process_dpi():
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

    @staticmethod
    def _screen_size():
        user32 = ctypes.windll.user32
        return [int(user32.GetSystemMetrics(0)), int(user32.GetSystemMetrics(1))]

    @staticmethod
    def _high_word_signed(value):
        return ctypes.c_short((int(value) >> 16) & 0xFFFF).value

    def _append_action(self, action):
        now = time.monotonic()
        if self._last_event_time is None:
            dt = 0.0
        else:
            dt = max(0.0, min(30.0, now - self._last_event_time))
        self._last_event_time = now
        action["dt"] = round(dt, 3)
        self.actions.append(action)

    def _sample_foreground_window(self, force=False):
        now = time.monotonic()
        if not force and now - self._last_window_sample < 0.08:
            return
        self._last_window_sample = now
        try:
            info = WindowsWindowHelper.get_foreground_info(action_index=len(self.actions))
        except Exception:
            return
        if WindowsWindowHelper.is_system_window(info):
            return
        key = ((info.get("title") or ""), (info.get("class") or ""))
        if key == self._last_window_key:
            return
        self._last_window_key = key
        WindowsWindowHelper.append_unique_window(self._window_history, info)

    def _mouse_proc(self, n_code, w_param, l_param):
        if n_code == self.HC_ACTION:
            msg = int(w_param)
            info = ctypes.cast(
                l_param,
                ctypes.POINTER(self.MSLLHOOKSTRUCT),
            ).contents
            x = int(info.pt.x)
            y = int(info.pt.y)

            mouse_events = {
                self.WM_LBUTTONDOWN: ("down", "left"),
                self.WM_LBUTTONUP: ("up", "left"),
                self.WM_RBUTTONDOWN: ("down", "right"),
                self.WM_RBUTTONUP: ("up", "right"),
                self.WM_MBUTTONDOWN: ("down", "middle"),
                self.WM_MBUTTONUP: ("up", "middle"),
            }

            if msg in mouse_events:
                event, button = mouse_events[msg]
                if event == "down":
                    self._pressed_buttons.add(button)
                elif event == "up":
                    self._pressed_buttons.discard(button)
                self._append_action(
                    {
                        "type": "mouse",
                        "event": event,
                        "button": button,
                        "x": x,
                        "y": y,
                    }
                )
            elif msg == self.WM_MOUSEWHEEL:
                self._append_action(
                    {
                        "type": "mouse",
                        "event": "wheel",
                        "x": x,
                        "y": y,
                        "delta": self._high_word_signed(info.mouseData),
                    }
                )
            elif msg == self.WM_MOUSEMOVE and self._pressed_buttons:
                self._append_action(
                    {
                        "type": "mouse",
                        "event": "move",
                        "x": x,
                        "y": y,
                    }
                )

        return ctypes.windll.user32.CallNextHookEx(self._mouse_hook, n_code, w_param, l_param)

    def _keyboard_proc(self, n_code, w_param, l_param):
        if n_code == self.HC_ACTION:
            msg = int(w_param)
            info = ctypes.cast(
                l_param,
                ctypes.POINTER(self.KBDLLHOOKSTRUCT),
            ).contents
            vk_code = int(info.vkCode)
            if vk_code == self.VK_F8:
                self.stop_event.set()
                return 1

            if msg in (self.WM_KEYDOWN, self.WM_SYSKEYDOWN, self.WM_KEYUP, self.WM_SYSKEYUP):
                self._append_action(
                    {
                        "type": "key",
                        "event": "down" if msg in (self.WM_KEYDOWN, self.WM_SYSKEYDOWN) else "up",
                        "vk": vk_code,
                        "scan": int(info.scanCode),
                        "flags": int(info.flags),
                    }
                )

        return ctypes.windll.user32.CallNextHookEx(self._keyboard_hook, n_code, w_param, l_param)

    def record(self):
        self._prepare_process_dpi()
        try:
            self._taskbar_rects = WindowsWindowHelper.get_taskbar_rects()
        except Exception:
            self._taskbar_rects = []
        user32 = ctypes.windll.user32
        kernel32 = ctypes.windll.kernel32
        user32.SetWindowsHookExW.argtypes = [
            ctypes.c_int,
            self._hook_proc_type,
            ctypes.c_void_p,
            wintypes.DWORD,
        ]
        user32.SetWindowsHookExW.restype = ctypes.c_void_p
        user32.CallNextHookEx.argtypes = [
            ctypes.c_void_p,
            ctypes.c_int,
            wintypes.WPARAM,
            wintypes.LPARAM,
        ]
        user32.CallNextHookEx.restype = ctypes.c_ssize_t
        user32.UnhookWindowsHookEx.argtypes = [ctypes.c_void_p]
        kernel32.GetModuleHandleW.restype = ctypes.c_void_p

        module_handle = kernel32.GetModuleHandleW(None)
        self._mouse_hook = user32.SetWindowsHookExW(
            self.WH_MOUSE_LL,
            self._mouse_proc_ref,
            module_handle,
            0,
        )
        self._keyboard_hook = user32.SetWindowsHookExW(
            self.WH_KEYBOARD_LL,
            self._keyboard_proc_ref,
            module_handle,
            0,
        )
        if not self._mouse_hook or not self._keyboard_hook:
            if self._mouse_hook:
                user32.UnhookWindowsHookEx(self._mouse_hook)
                self._mouse_hook = None
            if self._keyboard_hook:
                user32.UnhookWindowsHookEx(self._keyboard_hook)
                self._keyboard_hook = None
            raise RuntimeError("无法启动全局录制钩子，请尝试以普通权限重新打开本程序。")

        try:
            msg = wintypes.MSG()
            while not self.stop_event.is_set():
                while user32.PeekMessageW(ctypes.byref(msg), None, 0, 0, self.PM_REMOVE):
                    user32.TranslateMessage(ctypes.byref(msg))
                    user32.DispatchMessageW(ctypes.byref(msg))
                self._sample_foreground_window()
                time.sleep(0.01)
        finally:
            if self._mouse_hook:
                user32.UnhookWindowsHookEx(self._mouse_hook)
                self._mouse_hook = None
            if self._keyboard_hook:
                user32.UnhookWindowsHookEx(self._keyboard_hook)
                self._keyboard_hook = None

        self._sample_foreground_window(force=True)
        target_windows = WindowsWindowHelper.select_target_windows(
            self._window_history,
            self.actions,
            self._taskbar_rects,
        )
        return {
            "version": 2,
            "recorded_at": datetime.datetime.now().isoformat(timespec="seconds"),
            "stop_hotkey": "F8",
            "screen": self._screen_size(),
            "taskbar_rects": self._taskbar_rects,
            "window_history": self._window_history,
            "target_windows": target_windows,
            "actions": self.actions,
        }


class WindowsActionPlayer:
    MOUSEEVENTF_MOVE = 0x0001
    MOUSEEVENTF_LEFTDOWN = 0x0002
    MOUSEEVENTF_LEFTUP = 0x0004
    MOUSEEVENTF_RIGHTDOWN = 0x0008
    MOUSEEVENTF_RIGHTUP = 0x0010
    MOUSEEVENTF_MIDDLEDOWN = 0x0020
    MOUSEEVENTF_MIDDLEUP = 0x0040
    MOUSEEVENTF_WHEEL = 0x0800
    KEYEVENTF_EXTENDEDKEY = 0x0001
    KEYEVENTF_KEYUP = 0x0002

    def __init__(self, recording):
        if sys.platform != "win32":
            raise RuntimeError("自动回放仅支持 Windows。")
        if not recording or not recording.get("actions"):
            raise RuntimeError("还没有可用的导出 A 录制。")
        self.recording = recording
        self.user32 = ctypes.windll.user32
        self.target_context = None

    @staticmethod
    def _prepare_process_dpi():
        try:
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            pass

    def _scale_point(self, action):
        screen = self.recording.get("screen") or []
        recorded_w = float(screen[0]) if len(screen) > 0 and screen[0] else 0.0
        recorded_h = float(screen[1]) if len(screen) > 1 and screen[1] else 0.0
        current_w = float(self.user32.GetSystemMetrics(0))
        current_h = float(self.user32.GetSystemMetrics(1))
        sx = current_w / recorded_w if recorded_w else 1.0
        sy = current_h / recorded_h if recorded_h else 1.0
        x = int(round(float(action.get("x", 0)) * sx))
        y = int(round(float(action.get("y", 0)) * sy))
        x = max(0, min(int(current_w) - 1, x))
        y = max(0, min(int(current_h) - 1, y))
        return x, y

    def _point_for_action(self, action):
        try:
            raw_x = int(round(float(action.get("x", 0))))
            raw_y = int(round(float(action.get("y", 0))))
        except Exception:
            return self._scale_point(action)

        if self.target_context:
            recorded_rect = (self.target_context.get("recorded") or {}).get("rect")
            current_rect = (self.target_context.get("current") or {}).get("rect")
            if (
                recorded_rect
                and current_rect
                and len(recorded_rect) == 4
                and len(current_rect) == 4
                and WindowsWindowHelper.point_in_rect(raw_x, raw_y, recorded_rect)
            ):
                r_left, r_top, r_right, r_bottom = [float(v) for v in recorded_rect]
                c_left, c_top, c_right, c_bottom = [float(v) for v in current_rect]
                r_width = max(1.0, r_right - r_left)
                r_height = max(1.0, r_bottom - r_top)
                c_width = max(1.0, c_right - c_left)
                c_height = max(1.0, c_bottom - c_top)
                x = int(round(c_left + (raw_x - r_left) * c_width / r_width))
                y = int(round(c_top + (raw_y - r_top) * c_height / r_height))
                return x, y

        return self._scale_point(action)

    def _is_recorded_taskbar_action(self, action):
        taskbar_rects = self.recording.get("taskbar_rects") or []
        if taskbar_rects:
            return WindowsWindowHelper.action_in_rects(action, taskbar_rects)

        if action.get("type") != "mouse":
            return False
        screen = self.recording.get("screen") or []
        if len(screen) < 2 or not screen[1]:
            return False
        try:
            y = int(float(action.get("y", 0)))
        except Exception:
            return False
        return y >= int(screen[1]) - 90

    def replay(self):
        self._prepare_process_dpi()
        self.target_context = WindowsWindowHelper.activate_recording_target(self.recording)
        skip_leading_taskbar = bool(self.target_context)
        mouse_flags = {
            ("left", "down"): self.MOUSEEVENTF_LEFTDOWN,
            ("left", "up"): self.MOUSEEVENTF_LEFTUP,
            ("right", "down"): self.MOUSEEVENTF_RIGHTDOWN,
            ("right", "up"): self.MOUSEEVENTF_RIGHTUP,
            ("middle", "down"): self.MOUSEEVENTF_MIDDLEDOWN,
            ("middle", "up"): self.MOUSEEVENTF_MIDDLEUP,
        }

        for action in self.recording.get("actions", []):
            if skip_leading_taskbar and self._is_recorded_taskbar_action(action):
                continue
            skip_leading_taskbar = False
            time.sleep(max(0.0, min(30.0, float(action.get("dt", 0) or 0))))
            action_type = action.get("type")
            if action_type == "mouse":
                event = action.get("event")
                x, y = self._point_for_action(action)
                self.user32.SetCursorPos(x, y)
                if event == "move":
                    self.user32.mouse_event(self.MOUSEEVENTF_MOVE, 0, 0, 0, 0)
                elif event == "wheel":
                    self.user32.mouse_event(
                        self.MOUSEEVENTF_WHEEL,
                        0,
                        0,
                        int(action.get("delta", 0) or 0),
                        0,
                    )
                else:
                    flag = mouse_flags.get((action.get("button"), event))
                    if flag:
                        self.user32.mouse_event(flag, 0, 0, 0, 0)
            elif action_type == "key":
                flags = 0
                if int(action.get("flags", 0) or 0) & 0x01:
                    flags |= self.KEYEVENTF_EXTENDEDKEY
                if action.get("event") == "up":
                    flags |= self.KEYEVENTF_KEYUP
                self.user32.keybd_event(
                    int(action.get("vk", 0) or 0),
                    int(action.get("scan", 0) or 0),
                    flags,
                    0,
                )


class App:
    def __init__(self, root):
        self.root = root
        self.file_a = None
        self.file_b = None
        self.drop_zone_a = None
        self.drop_zone_b = None
        self.status_label = None
        self.process_btn = None
        self.clear_btn = None
        self.export_a_record_btn = None
        self.export_a_replay_btn = None
        self.export_a_status_label = None
        self.export_a_recording_thread = None
        self.export_a_replay_thread = None
        self.export_a_recording_stop_event = None
        self.row_height_var = tk.IntVar()
        self.row_height_text_var = tk.StringVar()
        self.ward_var = tk.StringVar()
        self.group_var = tk.StringVar()
        self.new_group_var = tk.StringVar()
        self.group_combo = None
        self.bed_vars = {}
        self.bed_buttons = {}
        self.bed_normal_font = None
        self.bed_selected_font = None
        self.maintenance_status_label = None
        cfg = load_config()
        self.row_height_var.set(max(20, min(80, int(cfg["row_height"]))))
        self.row_height_text_var.set(str(self.row_height_var.get()))
        self.ward_var.set(cfg["ward"])
        self.group_beds = cfg["groups"]
        self.group_var.set(cfg["current_group"])
        self.setup_ui()

    def normalize_row_height_value(self, value=None):
        if value is None:
            value = self.row_height_var.get()
        try:
            value = int(float(str(value).strip()))
        except Exception:
            try:
                value = int(self.row_height_var.get())
            except Exception:
                value = 30
        return max(20, min(80, value))

    def set_row_height(self, value=None, persist=True):
        value = self.normalize_row_height_value(value)
        if self.row_height_var.get() != value:
            self.row_height_var.set(value)
        if self.row_height_text_var.get() != str(value):
            self.row_height_text_var.set(str(value))
        if persist:
            save_config(row_height=value)
        return value

    def commit_row_height(self, *_):
        self.set_row_height(self.row_height_text_var.get(), persist=True)
        return True

    def setup_ui(self):
        self.root.title(APP_WINDOW_TITLE)
        self.root.geometry("1080x780")
        self.root.resizable(False, False)
        self.root.configure(bg="#F5F8FB")

        try:
            self.root.iconbitmap(get_resource_path("icon.ico"))
        except Exception:
            pass

        font_main = ("Microsoft YaHei UI", 10)
        font_bold = ("Microsoft YaHei UI", 10, "bold")
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure(
            "Workstation.TCombobox",
            fieldbackground="#FFFFFF",
            background="#FFFFFF",
            foreground="#07182F",
            bordercolor="#CBD7E3",
            arrowcolor="#07182F",
            padding=5,
        )

        def make_card(parent, bg="#FFFFFF", border="#D9E2EC"):
            outer = tk.Frame(parent, bg=border, bd=0)
            inner = tk.Frame(outer, bg=bg, bd=0)
            inner.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
            return outer, inner

        def make_label(parent, text, size=10, weight="normal", fg="#07182F", bg="#FFFFFF"):
            return tk.Label(
                parent,
                text=text,
                font=("Microsoft YaHei UI", size, weight),
                fg=fg,
                bg=bg,
                anchor=tk.W,
                justify=tk.LEFT,
            )

        shell = tk.Frame(self.root, bg="#F5F8FB")
        shell.pack(fill=tk.BOTH, expand=True)

        sidebar = tk.Frame(shell, bg="#123B57", width=250)
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)

        side_inner = tk.Frame(sidebar, bg="#123B57")
        side_inner.pack(fill=tk.BOTH, expand=True, padx=20, pady=24)

        tk.Label(
            side_inner,
            text="查房表\n更新工具",
            font=("Microsoft YaHei UI", 23, "bold"),
            fg="#FFFFFF",
            bg="#123B57",
            justify=tk.LEFT,
            anchor=tk.W,
        ).pack(fill=tk.X, pady=(0, 12))

        tk.Label(
            side_inner,
            text="把文件导入、床位维护、处理状态放在一个清晰工作台里。适合每天重复操作，扫一眼就知道下一步。",
            font=("Microsoft YaHei UI", 10),
            fg="#B9D3E5",
            bg="#123B57",
            justify=tk.LEFT,
            anchor=tk.W,
            wraplength=198,
        ).pack(fill=tk.X, pady=(0, 24))

        def stat_block(parent, value, caption):
            box = tk.Frame(parent, bg="#254E6A", highlightthickness=1, highlightbackground="#416B84")
            box.pack(fill=tk.X, pady=(0, 12))
            tk.Label(
                box,
                text=value,
                font=("Microsoft YaHei UI", 24, "bold"),
                fg="#FFFFFF",
                bg="#254E6A",
                anchor=tk.W,
            ).pack(fill=tk.X, padx=14, pady=(10, 0))
            tk.Label(
                box,
                text=caption,
                font=("Microsoft YaHei UI", 11),
                fg="#FFFFFF",
                bg="#254E6A",
                anchor=tk.W,
            ).pack(fill=tk.X, padx=14, pady=(0, 12))
            return box.winfo_children()[0]

        self.ward_stat_label = stat_block(side_inner, self.normalize_ward_value() or DEFAULT_WARD, "当前病区")
        self.bed_stat_label = stat_block(side_inner, str(len(self.get_selected_beds())), "已维护床位")
        self.height_stat_label = stat_block(side_inner, f"{self.row_height_var.get()} 磅", "最小行高")

        main = tk.Frame(shell, bg="#F5F8FB")
        main.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=28, pady=24)

        files_frame = tk.Frame(main, bg="#F5F8FB")
        files_frame.pack(fill=tk.X, pady=(0, 14))

        file_a_panel = tk.Frame(files_frame, bg="#F5F8FB")
        file_a_panel.pack(side=tk.LEFT, padx=(0, 16), anchor=tk.N)

        self.drop_zone_a = DropZone(
            file_a_panel,
            "病历系统导出列表\n单击选择，或把 Excel 拖到这里",
            color="#FFFFFF",
            zone_letter="A",
        )
        self.drop_zone_a.bind_select(lambda: self.select_file("A"))
        self.drop_zone_a.register_dnd(lambda e: self.on_drop(e, "A"))
        self.drop_zone_a.parent_app = self
        self.drop_zone_a.pack(side=tk.TOP)

        self.export_a_record_btn = tk.Button(
            self.drop_zone_a.header_frame,
            text="录制导出A",
            font=("Microsoft YaHei UI", 9, "bold"),
            bg="#E9EEF5",
            fg="#334155",
            activebackground="#DDE5EE",
            activeforeground="#07182F",
            relief=tk.FLAT,
            width=10,
            command=self.start_record_export_a,
        )
        self.export_a_replay_btn = tk.Button(
            self.drop_zone_a.header_frame,
            text="自动导出A",
            font=("Microsoft YaHei UI", 9, "bold"),
            bg="#EAF4FF",
            fg="#1D4ED8",
            activebackground="#D7E9FF",
            activeforeground="#1D4ED8",
            relief=tk.FLAT,
            width=10,
            command=self.start_replay_export_a,
        )
        self.export_a_replay_btn.pack(side=tk.RIGHT, ipady=2)
        self.export_a_record_btn.pack(side=tk.RIGHT, padx=(0, 8), ipady=2)

        file_b_panel = tk.Frame(files_frame, bg="#F5F8FB")
        file_b_panel.pack(side=tk.LEFT, anchor=tk.N)

        self.drop_zone_b = DropZone(
            file_b_panel,
            "昨日查房表\n用于保留格式并更新床位信息",
            color="#FFFFFF",
            zone_letter="B",
        )
        self.drop_zone_b.bind_select(lambda: self.select_file("B"))
        self.drop_zone_b.register_dnd(lambda e: self.on_drop(e, "B"))
        self.drop_zone_b.parent_app = self
        self.drop_zone_b.pack(side=tk.TOP)

        work_frame = tk.Frame(main, bg="#F5F8FB")
        work_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 18))

        settings_outer, settings = make_card(work_frame)
        settings_outer.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 18))
        settings_outer.configure(width=250)
        settings_outer.pack_propagate(False)
        settings.configure(padx=16, pady=14)

        make_label(settings, "基础设置", 12, "bold").pack(fill=tk.X, pady=(0, 8))

        make_label(settings, "病区", 10, "bold").pack(fill=tk.X, pady=(0, 4))
        ward_entry = tk.Entry(
            settings,
            textvariable=self.ward_var,
            width=8,
            justify=tk.CENTER,
            font=("Microsoft YaHei UI", 13, "bold"),
            relief=tk.SOLID,
            bd=1,
            highlightthickness=0,
        )
        ward_entry.pack(fill=tk.X, ipady=4)
        ward_entry.bind("<KeyRelease>", self.on_ward_changed)
        ward_entry.bind("<FocusOut>", self.on_ward_focus_out)

        make_label(settings, "最小行高", 10, "bold").pack(fill=tk.X, pady=(10, 4))
        row_height_frame = tk.Frame(settings, bg="#FFFFFF")
        row_height_frame.pack(fill=tk.X)
        row_height_slider = tk.Scale(
            row_height_frame,
            from_=20,
            to=80,
            orient=tk.HORIZONTAL,
            length=150,
            showvalue=0,
            sliderlength=16,
            troughcolor="#D8E2EC",
            bg="#FFFFFF",
            activebackground="#246BFE",
            bd=0,
            highlightthickness=0,
            variable=self.row_height_var,
            command=lambda v: save_config(row_height=int(float(v))),
        )
        row_height_slider.pack(side=tk.LEFT, fill=tk.X, expand=True)

        row_height_value_label = tk.Label(
            row_height_frame,
            text=f"{self.row_height_var.get()} 磅",
            font=("Microsoft YaHei UI", 11, "bold"),
            bg="#FFFFFF",
            fg="#246BFE",
            width=6,
            anchor=tk.E,
        )
        row_height_value_label.pack(side=tk.RIGHT, padx=(8, 0))

        def update_height_label(*_):
            value = f"{self.row_height_var.get()} 磅"
            row_height_value_label.config(text=value)
            if hasattr(self, "height_stat_label"):
                self.height_stat_label.config(text=value)

        self.row_height_var.trace_add("write", update_height_label)
        row_height_frame.pack_forget()

        compact_height_frame = tk.Frame(settings, bg="#FFFFFF")
        compact_height_frame.pack(fill=tk.X)

        def adjust_row_height(delta):
            value = self.normalize_row_height_value() + delta
            self.set_row_height(value, persist=True)

        tk.Button(
            compact_height_frame,
            text="-",
            font=("Microsoft YaHei UI", 12, "bold"),
            bg="#E9EEF5",
            fg="#334155",
            activebackground="#DDE5EE",
            relief=tk.FLAT,
            width=3,
            command=lambda: adjust_row_height(-1),
        ).pack(side=tk.LEFT, ipady=0)

        compact_height_value_entry = tk.Entry(
            compact_height_frame,
            textvariable=self.row_height_text_var,
            font=("Microsoft YaHei UI", 11, "bold"),
            bg="#F8FAFC",
            fg="#246BFE",
            width=9,
            justify=tk.CENTER,
            relief=tk.SOLID,
            bd=1,
        )
        compact_height_value_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=6, ipady=3)
        compact_height_value_entry.bind("<Return>", self.commit_row_height)
        compact_height_value_entry.bind("<FocusOut>", self.commit_row_height)

        tk.Button(
            compact_height_frame,
            text="+",
            font=("Microsoft YaHei UI", 12, "bold"),
            bg="#E9EEF5",
            fg="#334155",
            activebackground="#DDE5EE",
            relief=tk.FLAT,
            width=3,
            command=lambda: adjust_row_height(1),
        ).pack(side=tk.LEFT, ipady=0)

        def update_compact_height_label(*_):
            value = str(self.row_height_var.get())
            if self.row_height_text_var.get() != value:
                self.row_height_text_var.set(value)

        self.row_height_var.trace_add("write", update_compact_height_label)

        make_label(settings, "医疗组", 10, "bold").pack(fill=tk.X, pady=(10, 4))
        self.group_combo = ttk.Combobox(
            settings,
            textvariable=self.group_var,
            values=list(self.group_beds.keys()),
            state="readonly",
            width=14,
            font=font_main,
            style="Workstation.TCombobox",
        )
        self.group_combo.pack(fill=tk.X, ipady=3)
        self.group_combo.bind("<<ComboboxSelected>>", self.on_group_selected)

        tk.Entry(
            settings,
            textvariable=self.new_group_var,
            font=("Microsoft YaHei UI", 10),
            relief=tk.SOLID,
            bd=1,
        ).pack(fill=tk.X, pady=(6, 0), ipady=3)

        group_actions_frame = tk.Frame(settings, bg="#FFFFFF")
        group_actions_frame.pack(fill=tk.X, pady=(6, 0))
        tk.Button(
            group_actions_frame,
            text="新建",
            font=("Microsoft YaHei UI", 9, "bold"),
            bg="#E9EEF5",
            fg="#334155",
            activebackground="#DDE5EE",
            relief=tk.FLAT,
            command=self.create_medical_group,
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, ipady=2)
        tk.Button(
            group_actions_frame,
            text="改名",
            font=("Microsoft YaHei UI", 9, "bold"),
            bg="#E9EEF5",
            fg="#334155",
            activebackground="#DDE5EE",
            relief=tk.FLAT,
            command=self.rename_medical_group,
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0), ipady=2)
        tk.Button(
            group_actions_frame,
            text="删除",
            font=("Microsoft YaHei UI", 9, "bold"),
            bg="#FDEDEC",
            fg="#C0392B",
            activebackground="#FADBD8",
            relief=tk.FLAT,
            command=self.delete_medical_group,
        ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 0), ipady=2)

        bed_outer, bed_card = make_card(work_frame)
        bed_outer.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        bed_card.configure(padx=18, pady=18)

        bed_header = tk.Frame(bed_card, bg="#FFFFFF")
        bed_header.pack(fill=tk.X, pady=(0, 12))
        make_label(bed_header, "床位范围", 12, "bold").pack(side=tk.LEFT)
        tk.Button(
            bed_header,
            text="全选",
            font=("Microsoft YaHei UI", 9, "bold"),
            bg="#E9EEF5",
            fg="#334155",
            relief=tk.FLAT,
            command=self.select_all_beds,
        ).pack(side=tk.RIGHT, padx=(6, 0), ipadx=8, ipady=3)
        tk.Button(
            bed_header,
            text="清空",
            font=("Microsoft YaHei UI", 9, "bold"),
            bg="#E9EEF5",
            fg="#334155",
            relief=tk.FLAT,
            command=self.clear_bed_selection,
        ).pack(side=tk.RIGHT, ipadx=8, ipady=3)

        bed_select_frame = tk.Frame(bed_card, bg="#FFFFFF")
        bed_select_frame.pack(anchor=tk.NW)

        self.bed_normal_font = ("Consolas", 9, "bold")
        self.bed_selected_font = ("Consolas", 9, "bold")

        for i in range(1, 51):
            bed = f"{i:02d}"
            var = tk.BooleanVar(value=False)
            self.bed_vars[bed] = var
            cb = tk.Checkbutton(
                bed_select_frame,
                text=bed,
                variable=var,
                command=self.on_beds_changed,
                font=self.bed_normal_font,
                bg="#F8FAFC",
                activebackground="#E3F2FD",
                fg="#07182F",
                activeforeground="#07182F",
                selectcolor="#F8FAFC",
                indicatoron=False,
                relief=tk.SOLID,
                bd=1,
                width=4,
                padx=2,
                pady=4,
            )
            cb.grid(row=(i - 1) // 10, column=(i - 1) % 10, padx=3, pady=3, sticky=tk.W)
            self.bed_buttons[bed] = cb

        self.maintenance_status_label = tk.Label(
            bed_card,
            text="",
            font=("Microsoft YaHei UI", 10),
            bg="#FFFFFF",
            fg="#657084",
            anchor=tk.W,
            justify=tk.LEFT,
        )
        self.maintenance_status_label.pack(fill=tk.X, pady=(14, 0))

        self.apply_initial_bed_selection()
        self.update_maintenance_status()

        footer_outer, footer = make_card(main)
        footer_outer.pack(fill=tk.X)
        footer_outer.configure(height=118)
        footer_outer.pack_propagate(False)
        footer.configure(padx=18, pady=12)
        footer.columnconfigure(0, weight=1)
        footer.rowconfigure(0, minsize=38)
        footer.rowconfigure(1, minsize=50)

        self.status_label = tk.Label(
            footer,
            text="等待导入两份 Excel",
            font=("Microsoft YaHei UI", 11, "bold"),
            bg="#FFFFFF",
            fg="#07182F",
            anchor=tk.W,
            justify=tk.LEFT,
            wraplength=720,
            height=2,
        )
        self.status_label.grid(row=0, column=0, sticky=tk.EW, pady=(0, 8))

        button_frame = tk.Frame(footer, bg="#FFFFFF")
        button_frame.grid(row=1, column=0, sticky=tk.E)
        button_frame.configure(width=464, height=50)
        button_frame.grid_propagate(False)
        button_frame.columnconfigure(0, weight=1, uniform="footer_buttons")
        button_frame.columnconfigure(1, weight=1, uniform="footer_buttons")
        button_frame.rowconfigure(0, weight=1)

        self.clear_btn = tk.Button(
            button_frame,
            text="清除重选",
            font=("Microsoft YaHei UI", 12, "bold"),
            bg="#E9EEF5",
            fg="#334155",
            activebackground="#DDE5EE",
            activeforeground="#07182F",
            relief=tk.FLAT,
            width=10,
            padx=0,
            pady=10,
            command=self.clear_all,
        )
        self.clear_btn.grid(row=0, column=0, sticky=tk.NSEW, padx=(0, 8))

        self.process_btn = tk.Button(
            button_frame,
            text="开始处理",
            font=("Microsoft YaHei UI", 12, "bold"),
            bg="#AEBECD",
            fg="#FFFFFF",
            activebackground="#1F5FDB",
            activeforeground="#FFFFFF",
            relief=tk.FLAT,
            width=10,
            padx=0,
            pady=10,
            state=tk.DISABLED,
            command=self.start_processing,
        )
        self.process_btn.grid(row=0, column=1, sticky=tk.NSEW)

        self.update_button_state()
        self.update_status()
        self.update_export_a_buttons()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def current_group_name(self):
        name = self.group_var.get().strip()
        if not name:
            name = DEFAULT_GROUP
            self.group_var.set(name)
        if name not in self.group_beds:
            self.group_beds[name] = []
        return name

    def save_current_group_beds(self):
        self.group_beds[self.current_group_name()] = self.get_selected_beds()

    def persist_maintenance_config(self):
        save_config(
            ward=self.normalize_ward_value() or DEFAULT_WARD,
            beds=self.get_selected_beds(),
            groups=self.group_beds,
            current_group=self.current_group_name(),
        )

    def refresh_group_combo(self):
        if self.group_combo:
            self.group_combo.config(values=list(self.group_beds.keys()))

    def apply_group_bed_selection(self):
        if not self.bed_vars:
            return
        selected_beds = set(self.group_beds.get(self.current_group_name(), []))
        for bed, var in self.bed_vars.items():
            var.set(bed in selected_beds)
        self.update_bed_button_styles()

    def apply_initial_bed_selection(self):
        self.apply_group_bed_selection()

    def on_group_selected(self, _event=None):
        self.apply_group_bed_selection()
        self.persist_maintenance_config()
        self.update_maintenance_status()

    def create_medical_group(self):
        name = self.new_group_var.get().strip()
        if not name:
            messagebox.showwarning("提示", "请输入医疗组名称")
            return
        if name not in self.group_beds:
            self.group_beds[name] = []
            self.refresh_group_combo()
        self.group_var.set(name)
        self.new_group_var.set("")
        self.apply_group_bed_selection()
        self.persist_maintenance_config()
        self.update_maintenance_status()
        self.update_status()

    def rename_medical_group(self):
        old_name = self.current_group_name()
        new_name = self.new_group_var.get().strip()
        if not new_name:
            messagebox.showwarning("提示", "请输入新的医疗组名称")
            return
        if new_name == old_name:
            self.new_group_var.set("")
            return
        if new_name in self.group_beds:
            messagebox.showwarning("提示", "该医疗组名称已存在")
            return

        self.save_current_group_beds()
        updated_groups = {}
        for group_name, beds in self.group_beds.items():
            if group_name == old_name:
                updated_groups[new_name] = beds
            else:
                updated_groups[group_name] = beds
        self.group_beds = updated_groups
        self.group_var.set(new_name)
        self.new_group_var.set("")
        self.refresh_group_combo()
        self.apply_group_bed_selection()
        self.persist_maintenance_config()
        self.update_maintenance_status()
        self.update_status()

    def delete_medical_group(self):
        group_name = self.current_group_name()
        if len(self.group_beds) <= 1:
            messagebox.showwarning("提示", "至少需要保留一个医疗组")
            return
        if not messagebox.askyesno(
            "确认删除",
            f"确定删除医疗组“{group_name}”吗？\n该组维护的床位配置也会被删除。",
        ):
            return

        self.group_beds.pop(group_name, None)
        next_group = next(iter(self.group_beds), DEFAULT_GROUP)
        self.group_var.set(next_group)
        self.new_group_var.set("")
        self.refresh_group_combo()
        self.apply_group_bed_selection()
        self.persist_maintenance_config()
        self.update_maintenance_status()
        self.update_status()

    def normalize_ward_value(self, value=None):
        value = self.ward_var.get() if value is None else value
        digits = ''.join(ch for ch in str(value) if ch.isdigit())[:2]
        if not digits:
            return ""
        ward_num = int(digits)
        if ward_num < 1:
            ward_num = 1
        elif ward_num > 99:
            ward_num = 99
        return f"{ward_num:02d}"

    def on_ward_changed(self, _event=None):
        digits = ''.join(ch for ch in self.ward_var.get() if ch.isdigit())[:2]
        if digits != self.ward_var.get():
            self.ward_var.set(digits)
        ward = self.normalize_ward_value(digits)
        if ward:
            self.save_current_group_beds()
            self.persist_maintenance_config()
        self.update_maintenance_status()

    def on_ward_focus_out(self, _event=None):
        ward = self.normalize_ward_value()
        if not ward:
            ward = DEFAULT_WARD
        self.ward_var.set(ward)
        self.save_current_group_beds()
        self.persist_maintenance_config()
        self.update_maintenance_status()

    def get_selected_beds(self):
        if not self.bed_vars:
            return []
        return [bed for bed in sorted(self.bed_vars) if self.bed_vars[bed].get()]

    def update_bed_button_styles(self):
        for bed, button in self.bed_buttons.items():
            selected = self.bed_vars.get(bed) and self.bed_vars[bed].get()
            if selected:
                button.config(
                    bg="#246BFE",
                    fg="#FFFFFF",
                    activebackground="#1F5FDB",
                    activeforeground="#FFFFFF",
                    selectcolor="#246BFE",
                    relief=tk.SOLID,
                    font=self.bed_selected_font,
                )
            else:
                button.config(
                    bg="#F8FAFC",
                    fg="#07182F",
                    activebackground="#E3F2FD",
                    activeforeground="#07182F",
                    selectcolor="#F8FAFC",
                    relief=tk.SOLID,
                    font=self.bed_normal_font,
                )

    def on_beds_changed(self, _event=None):
        self.save_current_group_beds()
        self.persist_maintenance_config()
        self.update_bed_button_styles()
        self.update_maintenance_status()

    def select_all_beds(self):
        if not self.bed_vars:
            return
        for var in self.bed_vars.values():
            var.set(True)
        self.on_beds_changed()

    def clear_bed_selection(self):
        if not self.bed_vars:
            return
        for var in self.bed_vars.values():
            var.set(False)
        self.on_beds_changed()

    def get_maintained_bed_codes(self):
        ward = self.normalize_ward_value() or DEFAULT_WARD
        return [f"{ward}.{bed}" for bed in self.get_selected_beds()]

    def update_maintenance_status(self):
        if not self.maintenance_status_label:
            return
        ward = self.normalize_ward_value() or DEFAULT_WARD
        count = len(self.get_selected_beds())
        self.maintenance_status_label.config(text=f"{self.current_group_name()} | 病区 {ward} | 已选 {count} 张床")
        if hasattr(self, "ward_stat_label"):
            self.ward_stat_label.config(text=ward)
        if hasattr(self, "bed_stat_label"):
            self.bed_stat_label.config(text=str(count))

    def on_drop(self, event, zone):
        try:
            data = event.data
            if isinstance(data, str):
                files = self.parse_drop_data(data)
                if files:
                    file_path = files[0]
                    self.set_file(zone, file_path)
        except Exception as e:
            messagebox.showerror("错误", f"处理拖放失败: {str(e)}")

    def parse_drop_data(self, data):
        if sys.platform == 'win32':
            files = []
            parts = re.split(r'[{}\s]+', data)
            for part in parts:
                if part and (part.endswith('.xls') or part.endswith('.xlsx')) and os.path.isfile(part):
                    files.append(part)
                elif os.path.isfile(part):
                    ext = os.path.splitext(part)[1].lower()
                    if ext in ['.xls', '.xlsx']:
                        files.append(part)
            return files
        else:
            if os.path.isfile(data):
                return [data]
            return []

    def set_file(self, zone, path):
        if zone == 'A':
            self.file_a = path
            self.drop_zone_a.set_file(path)
        else:
            self.file_b = path
            self.drop_zone_b.set_file(path)
        self.update_button_state()
        self.update_status()

    def select_file(self, zone):
        from tkinter import filedialog
        filetypes = [("Excel 文件", "*.xls *.xlsx"), ("所有文件", "*.*")]
        title = (
            "选择文件 A：病历系统导出列表"
            if zone == "A"
            else "选择文件 B：昨日查房表"
        )
        path = filedialog.askopenfilename(title=title, filetypes=filetypes)
        if path:
            self.set_file(zone, path)

    def load_export_a_recording(self):
        try:
            if not os.path.exists(EXPORT_A_RECORDING_FILE):
                return None
            with open(EXPORT_A_RECORDING_FILE, "r", encoding="utf-8") as f:
                recording = json.load(f)
            actions = recording.get("actions") if isinstance(recording, dict) else None
            if not isinstance(actions, list) or not actions:
                return None
            return recording
        except Exception:
            return None

    def save_export_a_recording(self, recording):
        config_dir = os.path.dirname(EXPORT_A_RECORDING_FILE)
        if config_dir:
            os.makedirs(config_dir, exist_ok=True)
        with open(EXPORT_A_RECORDING_FILE, "w", encoding="utf-8") as f:
            json.dump(recording, f, ensure_ascii=False, indent=2)

    def has_export_a_recording(self):
        return self.load_export_a_recording() is not None

    def _is_thread_running(self, thread):
        return bool(thread and thread.is_alive())

    def _set_export_a_status(self, message, color="#657084"):
        if self.export_a_status_label:
            self.export_a_status_label.config(text=message, fg=color)

    def update_export_a_buttons(self):
        recording_active = self._is_thread_running(self.export_a_recording_thread)
        replay_active = self._is_thread_running(self.export_a_replay_thread)
        busy = recording_active or replay_active
        has_recording = self.has_export_a_recording()

        if self.export_a_record_btn:
            self.export_a_record_btn.config(
                text="录制中..." if recording_active else "录制导出A",
                state=tk.DISABLED if busy else tk.NORMAL,
            )
        if self.export_a_replay_btn:
            self.export_a_replay_btn.config(
                text="执行中..." if replay_active else "自动导出A",
                state=tk.NORMAL if has_recording and not busy else tk.DISABLED,
            )
        if not busy:
            if has_recording:
                self._set_export_a_status("已保存导出动作，可直接自动导入列表 A。", "#17835B")

    def start_record_export_a(self):
        if sys.platform != "win32":
            messagebox.showwarning("提示", "自动录制仅支持 Windows。")
            return
        if self._is_thread_running(self.export_a_recording_thread):
            return
        ok = messagebox.askokcancel(
            "录制导出A",
            "点击确定后，本窗口会最小化。\n"
            "请在医师工作站里完整操作一次“导出患者列表 A”。\n"
            "程序会记录目标窗口，之后任务栏图标位置变化也能识别。\n"
            "导出完成后按 F8 停止录制。",
        )
        if not ok:
            return

        self.export_a_recording_stop_event = threading.Event()
        self._set_export_a_status("录制准备中，完成导出后按 F8 停止。", "#246BFE")
        try:
            self.root.iconify()
        except tk.TclError:
            pass
        self.export_a_recording_thread = threading.Thread(
            target=self._record_export_a_worker,
            daemon=True,
        )
        self.export_a_recording_thread.start()
        self.update_export_a_buttons()

    def _record_export_a_worker(self):
        recording = None
        error = None
        try:
            time.sleep(0.8)
            recorder = WindowsActionRecorder(self.export_a_recording_stop_event)
            recording = recorder.record()
            if not recording.get("actions"):
                error = "没有录制到任何操作。"
            else:
                self.save_export_a_recording(recording)
        except Exception as exc:
            error = str(exc)

        def finish():
            self.export_a_recording_thread = None
            self.export_a_recording_stop_event = None
            try:
                self.root.deiconify()
                self.root.lift()
            except tk.TclError:
                pass
            if error:
                self._set_export_a_status(f"录制失败：{error}", "#C73B4A")
                messagebox.showerror("录制失败", error)
            else:
                count = len(recording.get("actions", [])) if recording else 0
                self._set_export_a_status(f"录制完成，已保存 {count} 个动作。", "#17835B")
                messagebox.showinfo("录制完成", "导出 A 的自动动作已保存。")
            self.update_export_a_buttons()

        try:
            self.root.after(0, finish)
        except tk.TclError:
            pass

    def start_replay_export_a(self):
        if sys.platform != "win32":
            messagebox.showwarning("提示", "自动回放仅支持 Windows。")
            return
        if self._is_thread_running(self.export_a_replay_thread):
            return

        recording = self.load_export_a_recording()
        if not recording:
            messagebox.showwarning("提示", "还没有录制导出 A 的动作，请先录制一次。")
            self.update_export_a_buttons()
            return

        self._set_export_a_status("正在自动导出列表 A，请稍候...", "#246BFE")
        try:
            self.root.iconify()
        except tk.TclError:
            pass
        self.export_a_replay_thread = threading.Thread(
            target=self._replay_export_a_worker,
            args=(recording,),
            daemon=True,
        )
        self.export_a_replay_thread.start()
        self.update_export_a_buttons()

    def _excel_search_dirs(self):
        candidates = [
            get_app_dir(),
            os.getcwd(),
            os.path.dirname(CONFIG_FILE),
        ]
        if self.file_a:
            candidates.append(os.path.dirname(self.file_a))
        user_home = os.path.expanduser("~")
        if user_home and os.path.isdir(user_home):
            candidates.extend(
                [
                    user_home,
                    os.path.join(user_home, "Desktop"),
                    os.path.join(user_home, "Downloads"),
                    os.path.join(user_home, "Documents"),
                    os.path.join(user_home, "桌面"),
                    os.path.join(user_home, "下载"),
                    os.path.join(user_home, "文档"),
                ]
            )

        result = []
        seen = set()
        for folder in candidates:
            if not folder:
                continue
            try:
                full = os.path.abspath(folder)
            except Exception:
                continue
            key = os.path.normcase(full)
            if key in seen or not os.path.isdir(full):
                continue
            seen.add(key)
            result.append(full)
        return result

    def _snapshot_excel_files(self):
        snapshot = {}
        for folder in self._excel_search_dirs():
            try:
                names = os.listdir(folder)
            except Exception:
                continue
            for name in names:
                if name.startswith("~$"):
                    continue
                ext = os.path.splitext(name)[1].lower()
                if ext not in (".xls", ".xlsx"):
                    continue
                path = os.path.join(folder, name)
                try:
                    snapshot[os.path.abspath(path)] = (
                        os.path.getmtime(path),
                        os.path.getsize(path),
                    )
                except OSError:
                    continue
        return snapshot

    def _is_excel_file_stable(self, path):
        try:
            size_1 = os.path.getsize(path)
            mtime_1 = os.path.getmtime(path)
            time.sleep(0.35)
            return size_1 == os.path.getsize(path) and mtime_1 == os.path.getmtime(path)
        except OSError:
            return False

    def _find_new_excel_file(self, before_snapshot, started_at):
        after_snapshot = self._snapshot_excel_files()
        candidates = []
        for path, signature in after_snapshot.items():
            old_signature = before_snapshot.get(path)
            if old_signature == signature:
                continue
            mtime, _size = signature
            if mtime >= started_at - 3:
                candidates.append((mtime, path))
        if not candidates:
            return None
        candidates.sort(reverse=True)
        return candidates[0][1]

    def _wait_for_new_excel_file(self, before_snapshot, started_at, timeout=25):
        deadline = time.time() + timeout
        latest_path = None
        while time.time() < deadline:
            latest_path = self._find_new_excel_file(before_snapshot, started_at)
            if latest_path and self._is_excel_file_stable(latest_path):
                return latest_path
            time.sleep(0.5)
        return latest_path

    def _replay_export_a_worker(self, recording):
        found_path = None
        error = None
        started_at = time.time()
        before_snapshot = self._snapshot_excel_files()
        try:
            time.sleep(0.8)
            WindowsActionPlayer(recording).replay()
            found_path = self._wait_for_new_excel_file(before_snapshot, started_at)
        except Exception as exc:
            error = str(exc)

        def finish():
            self.export_a_replay_thread = None
            try:
                self.root.deiconify()
                self.root.lift()
            except tk.TclError:
                pass
            if error:
                self._set_export_a_status(f"自动导出失败：{error}", "#C73B4A")
                messagebox.showerror("自动导出失败", error)
            elif found_path:
                self.set_file("A", found_path)
                self._set_export_a_status(
                    f"已自动导入列表 A：{os.path.basename(found_path)}",
                    "#17835B",
                )
            else:
                self._set_export_a_status("已执行动作，但未发现新 Excel。请手动选择列表 A。", "#A76400")
                messagebox.showwarning(
                    "未找到导出文件",
                    "自动动作已执行，但没有在常用位置发现新的 Excel 文件。\n"
                    "如果文件已导出，请手动选择列表 A。",
                )
            self.update_export_a_buttons()

        try:
            self.root.after(0, finish)
        except tk.TclError:
            pass

    def on_close(self):
        if self.export_a_recording_stop_event:
            self.export_a_recording_stop_event.set()
        try:
            self.root.destroy()
        except tk.TclError:
            pass

    def update_button_state(self):
        ready = bool(self.file_a and self.file_b)
        if ready:
            self.process_btn.config(state=tk.NORMAL, bg="#246BFE", activebackground="#1F5FDB")
        else:
            self.process_btn.config(state=tk.DISABLED, bg="#AEBECD", activebackground="#AEBECD")

    def update_status(self):
        if self.file_a and self.file_b:
            self.status_label.config(
                text=f"已就绪：{os.path.basename(self.file_a)} + {os.path.basename(self.file_b)}",
                fg="#17835B",
            )
        elif self.file_a:
            self.status_label.config(
                text=f"已选择文件 A：{os.path.basename(self.file_a)}，请继续导入文件 B。",
                fg="#A76400",
            )
        elif self.file_b:
            self.status_label.config(
                text=f"已选择文件 B：{os.path.basename(self.file_b)}，请继续导入文件 A。",
                fg="#A76400",
            )
        else:
            ward = self.normalize_ward_value() or DEFAULT_WARD
            count = len(self.get_selected_beds())
            self.status_label.config(
                text=f"等待导入两份 Excel。会输出 {ward} 病区，{self.current_group_name()}，{count} 张床位的更新表。",
                fg="#07182F",
            )

    def clear_all(self):
        self.file_a = None
        self.file_b = None
        self.drop_zone_a.clear()
        self.drop_zone_b.clear()
        self.update_button_state()
        self.update_status()

    def start_processing(self):
        self.commit_row_height()
        self.process_btn.config(state=tk.DISABLED, bg="#AEBECD")
        self.status_label.config(text="处理中，请稍候...", fg="#246BFE")
        self.root.update()

        try:
            path_a, path_b = self.identify_files(self.file_a, self.file_b)
            result = self.process_files(path_a, path_b)

            if result:
                self.status_label.config(text=f"处理完成，已保存：{os.path.basename(result)}", fg="#17835B")
                os.startfile(result)
            else:
                self.status_label.config(text="处理失败，请检查文件格式。", fg="#C73B4A")
        except Exception as e:
            self.status_label.config(text="处理失败", fg="#C73B4A")
            messagebox.showerror("错误", f"处理文件时发生错误：\n{str(e)}")
        finally:
            self.update_button_state()

    def identify_files(self, path_a, path_b):
        import xlrd

        file_a_has_status = False
        file_b_has_status = False
        
        for path, label in [(path_a, 'A'), (path_b, 'B')]:
            try:
                wb = xlrd.open_workbook(path)
                ws = wb.sheet_by_index(0)
                if ws.nrows > 0:
                    first_row = [str(ws.cell_value(0, col)).strip() for col in range(ws.ncols)]
                    if any("病历状态" in cell for cell in first_row):
                        file_a_has_status = True if label == 'A' else file_b_has_status
                        if label == 'A':
                            file_a_has_status = True
                        else:
                            file_b_has_status = True
            except Exception as e:
                raise Exception(f"读取文件{label}失败: {str(e)}")
        
        wb_a = xlrd.open_workbook(path_a)
        ws_a = wb_a.sheet_by_index(0)
        first_row_a = [str(ws_a.cell_value(0, col)).strip() for col in range(ws_a.ncols)]
        a_has_status = any("病历状态" in cell for cell in first_row_a)
        
        wb_b = xlrd.open_workbook(path_b)
        ws_b = wb_b.sheet_by_index(0)
        first_row_b = [str(ws_b.cell_value(0, col)).strip() for col in range(ws_b.ncols)]
        b_has_status = any("病历状态" in cell for cell in first_row_b)
        
        if a_has_status and not b_has_status:
            return path_a, path_b
        elif b_has_status and not a_has_status:
            return path_b, path_a
        elif a_has_status and b_has_status:
            raise Exception("错误：两个文件都包含'病历状态'列，无法区分文档A和B")
        else:
            raise Exception("错误：两个文件都不包含'病历状态'列，无法识别文档A")

    def find_column_row(self, ws, required_cols, max_rows=10):
        nrows = ws.nrows
        search_limit = min(nrows, max_rows)
        
        col_patterns = {
            'bed': ['床号', '床位', '床位号'],
            'name': ['姓名', '病人', '病人姓名'],
            'gender': ['性别', '男/女'],
            'age': ['年龄', '岁', '患者年龄'],
            'diagnosis': ['诊断', '病史', '临床诊断'],
            'status': ['病历状态', '状态']
        }
        
        for row_idx in range(search_limit):
            row_values = [str(ws.cell_value(row_idx, col)).strip() for col in range(ws.ncols)]
            found_cols = {}
            
            for col_idx, cell_value in enumerate(row_values):
                cell_lower = cell_value.lower()
                
                for col_name, patterns in col_patterns.items():
                    if col_name not in found_cols:
                        for pattern in patterns:
                            if pattern in cell_value or pattern.lower() in cell_lower:
                                found_cols[col_name] = col_idx
                                break
                        if col_name in found_cols:
                            break
            
            if all(col in found_cols for col in required_cols):
                return row_idx, found_cols
        
        return -1, {}

    def normalize_name(self, name):
        if name is None:
            return ""
        name_str = str(name).strip()
        name_str = name_str.replace(" ", "").replace("　", "")
        return name_str

    def normalize_bed_code(self, bed, default_ward=None):
        if bed is None:
            return ""
        digits = ''.join(ch for ch in str(bed).strip() if ch.isdigit())
        if not digits:
            return ""
        if len(digits) >= 4:
            return digits[-4:]
        if len(digits) == 3:
            return f"0{digits}"
        bed_num = int(digits)
        if default_ward and 1 <= bed_num <= 50:
            return f"{default_ward}{bed_num:02d}"
        return f"{bed_num:02d}"

    def bed_sort_key(self, bed, default_ward=None):
        text = "" if bed is None else str(bed).strip()
        text_lower = text.lower()
        has_zr = "zr" in text_lower
        normalized = self.normalize_bed_code(text, default_ward)
        if normalized:
            try:
                numeric = int(normalized)
            except ValueError:
                numeric = 999999
        else:
            numeric = 999999
        return (1 if has_zr else 0, numeric, text_lower)

    def _drop_b_only_duplicate_beds(self, rows, from_a_flags, bed_col_idx, source_rows):
        """同一床号多行且其中至少一行来自 A 时，删除仍为原 B、未被 A 匹配的行（旧患者出院、新患者占床）。"""
        if not rows or bed_col_idx < 0:
            return rows, from_a_flags, source_rows

        bed_to_indices = defaultdict(list)
        for i, row in enumerate(rows):
            if len(row) <= bed_col_idx:
                continue
            bed = self.normalize_bed_code(row[bed_col_idx], self.normalize_ward_value() or DEFAULT_WARD)
            if bed:
                bed_to_indices[bed].append(i)

        remove_idx = set()
        for indices in bed_to_indices.values():
            if len(indices) <= 1:
                continue
            if not any(from_a_flags[i] for i in indices):
                continue
            for i in indices:
                if not from_a_flags[i]:
                    remove_idx.add(i)

        if not remove_idx:
            return rows, from_a_flags, source_rows
        return (
            [row for i, row in enumerate(rows) if i not in remove_idx],
            [flag for i, flag in enumerate(from_a_flags) if i not in remove_idx],
            [source for i, source in enumerate(source_rows) if i not in remove_idx],
        )

    def _drop_b_rows_for_empty_a_beds(self, rows, from_a_flags, source_rows, bed_col_idx, empty_beds):
        """A 表明确出现空床时，删除 B 表中该床位未被 A 匹配到的旧患者记录。"""
        if not rows or bed_col_idx < 0 or not empty_beds:
            return rows, from_a_flags, source_rows

        kept_rows = []
        kept_flags = []
        kept_sources = []
        for row, from_a, source_row in zip(rows, from_a_flags, source_rows):
            bed = ""
            if len(row) > bed_col_idx:
                bed = self.normalize_bed_code(row[bed_col_idx], self.normalize_ward_value() or DEFAULT_WARD)
            if bed in empty_beds and not from_a:
                continue
            kept_rows.append(row)
            kept_flags.append(from_a)
            kept_sources.append(source_row)
        return kept_rows, kept_flags, kept_sources

    def _append_missing_maintained_beds(self, rows, from_a_flags, source_rows, bed_col_idx, col_count, maintained_beds):
        if not maintained_beds or bed_col_idx < 0:
            return rows, from_a_flags, source_rows

        ward = self.normalize_ward_value() or DEFAULT_WARD
        existing_beds = set()
        for row in rows:
            if len(row) > bed_col_idx:
                bed = self.normalize_bed_code(row[bed_col_idx], ward)
                if bed:
                    existing_beds.add(bed)

        for bed_display in maintained_beds:
            bed_code = self.normalize_bed_code(bed_display, ward)
            if not bed_code or bed_code in existing_beds:
                continue
            new_row = [""] * col_count
            new_row[bed_col_idx] = bed_display
            rows.append(new_row)
            from_a_flags.append(False)
            source_rows.append(None)
            existing_beds.add(bed_code)
        return rows, from_a_flags, source_rows

    def _text_width_units(self, text):
        width = 0.0
        for ch in text:
            if ch == "\t":
                width += 4.0
            elif ch.isspace():
                width += 0.5
            elif unicodedata.east_asian_width(ch) in ("W", "F"):
                width += 1.65
            elif unicodedata.east_asian_width(ch) == "A":
                width += 1.2
            elif ch in ".,;:!?()[]{}<>+-=*/\\'\"":
                width += 0.6
            else:
                width += 0.9
        return width

    def _font_cache_key(self, font_name, font_size):
        try:
            size = max(1, int(round(float(font_size or 11))))
        except Exception:
            size = 11
        return (font_name or "Microsoft YaHei UI", size)

    def _get_measure_font(self, font_name, font_size):
        if not hasattr(self, "_measure_font_cache"):
            self._measure_font_cache = {}
        key = self._font_cache_key(font_name, font_size)
        if key not in self._measure_font_cache:
            self._measure_font_cache[key] = tkfont.Font(
                family=key[0],
                size=key[1],
            )
        return self._measure_font_cache[key]

    def _column_width_pixels(self, column_width_chars):
        try:
            width = max(1.0, float(column_width_chars or 8.43))
        except Exception:
            width = 8.43
        if width < 1:
            return max(8, int(width * 12))
        return max(12, int(math.floor(width * 7 + 5)))

    def _measure_text_pixels(self, text, font_name, font_size):
        font = self._get_measure_font(font_name, font_size)
        return max(0, int(font.measure(text)))

    def _estimate_wrapped_lines_pixels(self, text, max_pixels, font_name, font_size):
        if not text:
            return 1
        line_count = 1
        current_width = 0
        for ch in text:
            char_text = "    " if ch == "\t" else ch
            try:
                char_width = max(1, self._measure_text_pixels(char_text, font_name, font_size))
            except Exception:
                return max(
                    1,
                    int(math.ceil(self._text_width_units(text) / max(3.6, float(max_pixels) / 7.0))),
                )
            if current_width > 0 and current_width + char_width > max_pixels:
                line_count += 1
                current_width = char_width
            else:
                current_width += char_width
        return max(1, line_count)

    def _estimate_cell_lines(self, value, column_width_chars, font_name=None, font_size=11):
        if value is None:
            return 1
        text = str(value)
        if not text:
            return 1

        usable_pixels = self._column_width_pixels(column_width_chars)
        line_count = 0
        for part in text.replace("\r\n", "\n").replace("\r", "\n").split("\n"):
            line_count += self._estimate_wrapped_lines_pixels(
                part,
                usable_pixels,
                font_name,
                font_size,
            )
        return max(1, line_count)

    def _estimate_line_height_points(self, font_name, font_size):
        try:
            size = max(1.0, float(font_size or 11))
        except Exception:
            size = 11.0
        try:
            font = self._get_measure_font(font_name, size)
            pixels_per_inch = float(self.root.winfo_fpixels("1i"))
            measured_points = float(font.metrics("linespace")) * 72.0 / pixels_per_inch
            return max(size + 4.0, measured_points + 1.0)
        except Exception:
            return max(size + 4.0, size * 1.25)

    def _estimate_row_height(
        self,
        row_data,
        col_count,
        get_column_width_chars,
        get_font_size=None,
        get_font_name=None,
    ):
        base_height = self.normalize_row_height_value()
        max_height = base_height
        for col_idx in range(col_count):
            value = row_data[col_idx] if col_idx < len(row_data) else ""
            font_size = 11
            if get_font_size:
                try:
                    font_size = float(get_font_size(col_idx) or 11)
                except Exception:
                    font_size = 11
            font_name = None
            if get_font_name:
                try:
                    font_name = get_font_name(col_idx)
                except Exception:
                    font_name = None
            lines = self._estimate_cell_lines(
                value,
                get_column_width_chars(col_idx),
                font_name,
                font_size,
            )
            line_height = self._estimate_line_height_points(font_name, font_size)
            content_height = lines * line_height
            if content_height > max_height:
                max_height = content_height

        return min(409, math.ceil(max_height * 4) / 4.0)

    def process_files(self, a_path, b_path):
        import xlrd

        wb_a = xlrd.open_workbook(a_path, formatting_info=True)
        ws_a = wb_a.sheet_by_index(0)
        
        wb_b = xlrd.open_workbook(b_path, formatting_info=True)
        ws_b = wb_b.sheet_by_index(0)
        
        required_cols_a = ['bed', 'name', 'gender', 'age', 'diagnosis', 'status']
        col_row_a, cols_a = self.find_column_row(ws_a, required_cols_a, max_rows=10)
        
        if col_row_a == -1:
            raise Exception("未能在文档A中找到必要的列（床号、姓名、性别、年龄、诊断、病历状态）")
        
        required_cols_b = ['bed', 'name', 'gender', 'age', 'diagnosis']
        col_row_b, cols_b = self.find_column_row(ws_b, required_cols_b, max_rows=6)
        
        if col_row_b == -1:
            raise Exception("未能在文档B中找到必要的列（床号、姓名、性别、年龄、诊断）")
        
        data_rows_a = []
        empty_beds_a = set()
        for row_idx in range(col_row_a + 1, ws_a.nrows):
            bed = str(ws_a.cell_value(row_idx, cols_a['bed'])).strip()
            bed_code = self.normalize_bed_code(bed, self.normalize_ward_value() or DEFAULT_WARD)
            name = ws_a.cell_value(row_idx, cols_a['name'])
            if name and str(name).strip():
                row_data = {
                    'bed': bed,
                    'name': self.normalize_name(name),
                    'original_name': str(name).strip(),
                    'gender': str(ws_a.cell_value(row_idx, cols_a['gender'])).strip(),
                    'age': str(ws_a.cell_value(row_idx, cols_a['age'])).strip(),
                    'diagnosis': str(ws_a.cell_value(row_idx, cols_a['diagnosis'])).strip()
                }
                data_rows_a.append(row_data)
            elif bed_code:
                empty_beds_a.add(bed_code)
        
        header_rows_b = []
        for row_idx in range(col_row_b + 1):
            row_data = []
            for col_idx in range(ws_b.ncols):
                cell = ws_b.cell(row_idx, col_idx)
                row_data.append(cell.value)
            header_rows_b.append(row_data)
        
        data_rows_b = []
        data_source_rows_b = []
        for row_idx in range(col_row_b + 1, ws_b.nrows):
            row_data = []
            for col_idx in range(ws_b.ncols):
                cell = ws_b.cell(row_idx, col_idx)
                row_data.append(cell.value)
            data_rows_b.append(row_data)
            data_source_rows_b.append(row_idx)
        
        b_col_count = ws_b.ncols
        
        b_lookup = {}
        for idx, row in enumerate(data_rows_b):
            if len(row) > cols_b['name']:
                name = self.normalize_name(row[cols_b['name']])
                if name:
                    b_lookup[name] = (idx, row)

        initial_b_count = len(data_rows_b)
        updated_b_indices = set()

        for a_row in data_rows_a:
            a_name = a_row['name']

            if a_name in b_lookup:
                row_idx, row = b_lookup[a_name]
                # 匹配成功后，将床号和病人核心字段全部从 A 表覆盖
                row[cols_b['bed']] = a_row['bed']
                row[cols_b['name']] = a_row['original_name']
                row[cols_b['gender']] = a_row['gender']
                row[cols_b['age']] = a_row['age']
                row[cols_b['diagnosis']] = a_row['diagnosis']
                b_lookup[a_name] = (row_idx, row)
                updated_b_indices.add(row_idx)
            else:
                new_row = [""] * b_col_count
                new_row[cols_b['bed']] = a_row['bed']
                new_row[cols_b['name']] = a_row['original_name']
                new_row[cols_b['gender']] = a_row['gender']
                new_row[cols_b['age']] = a_row['age']
                new_row[cols_b['diagnosis']] = a_row['diagnosis']
                data_rows_b.append(new_row)
                data_source_rows_b.append(None)

        # 与 data_rows_b 下标对齐：True=已由 A 覆盖或来自 A 追加，False=仍为原 B 行且本次未被 A 匹配
        row_from_a = [
            (i >= initial_b_count or i in updated_b_indices) for i in range(len(data_rows_b))
        ]

        def bed_sort_key(pair):
            row = pair[0]
            if len(row) > cols_b['bed']:
                return self.bed_sort_key(row[cols_b['bed']], self.normalize_ward_value() or DEFAULT_WARD)
            return (0, 999999, "")

        paired = list(zip(data_rows_b, row_from_a, data_source_rows_b))
        paired.sort(key=bed_sort_key)
        data_rows_b = [p[0] for p in paired]
        row_from_a = [p[1] for p in paired]
        data_source_rows_b = [p[2] for p in paired]

        # A 表中有床号但姓名为空，表示该床当前无患者；删除 B 表里这个床位的旧患者记录
        data_rows_b, row_from_a, data_source_rows_b = self._drop_b_rows_for_empty_a_beds(
            data_rows_b,
            row_from_a,
            data_source_rows_b,
            cols_b['bed'],
            empty_beds_a,
        )

        # 排序后、保存前：同一床号出现多行时，说明原 B 中占床患者已出院，新患者在 A 中；删掉仍为「仅 B」的旧行
        data_rows_b, row_from_a, data_source_rows_b = self._drop_b_only_duplicate_beds(
            data_rows_b,
            row_from_a,
            cols_b['bed'],
            data_source_rows_b,
        )

        maintained_beds = self.get_maintained_bed_codes()
        data_rows_b, row_from_a, data_source_rows_b = self._append_missing_maintained_beds(
            data_rows_b,
            row_from_a,
            data_source_rows_b,
            cols_b['bed'],
            b_col_count,
            maintained_beds,
        )
        final_rows = list(zip(data_rows_b, data_source_rows_b))
        final_rows.sort(
            key=lambda row: self.bed_sort_key(
                row[0][cols_b['bed']] if len(row[0]) > cols_b['bed'] else "",
                self.normalize_ward_value() or DEFAULT_WARD,
            )
        )
        data_rows_b = [row for row, _source in final_rows]
        data_source_rows_b = [source for _row, source in final_rows]

        output_path = self.save_output(data_rows_b, header_rows_b, b_col_count, b_path, ws_b, cols_b, data_source_rows_b)
        
        return output_path

    def save_output(self, data_rows, header_rows, col_count, b_path, ws_b_original, cols_b, data_source_rows=None):
        """B 为 .xlsx 时用 openpyxl 保留表头合并；B 为 .xls 时用 xlwt（openpyxl 不支持 .xls）。"""
        ext = os.path.splitext(b_path)[1].lower()
        if data_source_rows is None:
            data_source_rows = [None] * len(data_rows)
        if ext == ".xlsx":
            return self._save_output_xlsx(data_rows, header_rows, col_count, b_path, data_source_rows)
        return self._save_output_xls(data_rows, header_rows, col_count, b_path, ws_b_original, data_source_rows)

    def _save_output_xlsx(self, data_rows, header_rows, col_count, b_path, data_source_rows=None):
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        from openpyxl.utils import get_column_letter

        wb = load_workbook(b_path, rich_text=True)
        ws = wb.active
        header_row_count = len(header_rows)
        if data_source_rows is None:
            data_source_rows = [None] * len(data_rows)

        data_style_rows = []
        for source_row in range(header_row_count + 1, ws.max_row + 1):
            style_row = []
            for col_idx in range(col_count):
                source_cell = ws.cell(row=source_row, column=col_idx + 1)
                style_row.append({
                    "style": copy(source_cell._style),
                    "alignment": copy(source_cell.alignment),
                    "number_format": source_cell.number_format,
                    "protection": copy(source_cell.protection),
                    "value": copy(source_cell.value),
                    "font_name": source_cell.font.name or "Microsoft YaHei UI",
                    "font_size": source_cell.font.sz or 11,
                })
            data_style_rows.append(style_row)
        template_style_row = data_style_rows[0] if data_style_rows else None

        while ws.max_row > header_row_count:
            ws.delete_rows(header_row_count + 1)

        def _xlsx_col_width_chars(col_idx):
            column_letter = get_column_letter(col_idx + 1)
            width = ws.column_dimensions[column_letter].width
            return width if width else 8.43

        for row_idx, row_data in enumerate(data_rows):
            target_row = header_row_count + row_idx + 1
            source_row_idx = data_source_rows[row_idx] if row_idx < len(data_source_rows) else None
            row_source_styles = None
            if source_row_idx is not None:
                style_idx = source_row_idx - header_row_count
                if 0 <= style_idx < len(data_style_rows):
                    row_source_styles = data_style_rows[style_idx]
            if row_source_styles is None:
                row_source_styles = template_style_row
            for col_idx, value in enumerate(row_data):
                if col_idx < col_count:
                    cell = ws.cell(row=target_row, column=col_idx + 1)
                    style_info = None
                    if row_source_styles and col_idx < len(row_source_styles):
                        style_info = row_source_styles[col_idx]
                        cell._style = copy(style_info["style"])
                        cell.alignment = copy(style_info["alignment"])
                        cell.number_format = style_info["number_format"]
                        cell.protection = copy(style_info["protection"])
                    source_value = style_info["value"] if style_info else None
                    if type(source_value).__name__ == "CellRichText" and str(source_value) == str(value):
                        cell.value = copy(source_value)
                    else:
                        cell.value = value
                    cell.alignment = copy(cell.alignment)
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        text_rotation=cell.alignment.text_rotation,
                        wrap_text=True,
                        shrink_to_fit=cell.alignment.shrink_to_fit,
                        indent=cell.alignment.indent,
                    )
            estimated_height = self._estimate_row_height(
                row_data,
                col_count,
                _xlsx_col_width_chars,
                lambda col_idx, styles=row_source_styles: (
                    styles[col_idx]["font_size"] if styles and col_idx < len(styles) else 11
                ),
                lambda col_idx, styles=row_source_styles: (
                    styles[col_idx]["font_name"] if styles and col_idx < len(styles) else None
                ),
            )
            ws.row_dimensions[target_row].height = estimated_height
        # 打印区域：与表同宽，覆盖全部数据行
        last_data_row = header_row_count + len(data_rows)
        last_col_letter = get_column_letter(max(1, col_count - 1))
        ws.print_area = f"A1:{last_col_letter}{last_data_row}"
        # 页面设置：横向、A4、无页眉页脚，上下顶格
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.oddHeader.left.text = ""
        ws.oddHeader.center.text = ""
        ws.oddHeader.right.text = ""
        ws.oddFooter.left.text = ""
        ws.oddFooter.center.text = ""
        ws.oddFooter.right.text = ""
        ws.evenHeader.left.text = ""
        ws.evenHeader.center.text = ""
        ws.evenHeader.right.text = ""
        ws.evenFooter.left.text = ""
        ws.evenFooter.center.text = ""
        ws.evenFooter.right.text = ""
        ws.firstHeader.left.text = ""
        ws.firstHeader.center.text = ""
        ws.firstHeader.right.text = ""
        ws.firstFooter.left.text = ""
        ws.firstFooter.center.text = ""
        ws.firstFooter.right.text = ""
        ws.page_margins.top = 0
        ws.page_margins.bottom = 0
        ws.page_margins.header = 0
        ws.page_margins.footer = 0

        base_name = os.path.splitext(os.path.basename(b_path))[0]
        date_str = datetime.datetime.now().strftime("%Y%m%d")
        output_filename = f"{base_name}_{date_str}更新.xlsx"
        output_path = os.path.join(os.path.dirname(b_path), output_filename)
        wb.save(output_path)
        return output_path

    def _copy_xls_palette(self, src_book, dst_book):
        """复制 .xls 的自定义调色板，避免颜色索引在新工作簿里指向错误颜色。"""
        try:
            for colour_index, rgb in getattr(src_book, "colour_map", {}).items():
                if rgb is None or not (8 <= int(colour_index) <= 63):
                    continue
                red, green, blue = rgb
                dst_book.set_colour_RGB(int(colour_index), int(red), int(green), int(blue))
        except Exception:
            pass

    def _save_output_xls(self, data_rows, header_rows, col_count, b_path, ws_src, data_source_rows=None):
        """老版 .xls：用 xlwt 写出；表头区完全按 B 表合并（仅保留整段落在表头内的合并）。"""
        import xlwt

        wb = xlwt.Workbook(encoding="utf-8")
        self._copy_xls_palette(ws_src.book, wb)
        ws = wb.add_sheet("Sheet1")
        n_header = len(header_rows)
        if data_source_rows is None:
            data_source_rows = [None] * len(data_rows)

        header_alignment = xlwt.Alignment()
        header_alignment.horz = xlwt.Alignment.HORZ_CENTER
        header_alignment.vert = xlwt.Alignment.VERT_CENTER
        header_alignment.wrap = xlwt.Alignment.WRAP_AT_RIGHT
        header_style = xlwt.XFStyle()
        header_style.font = xlwt.Font()
        header_style.font.height = 220
        header_style.alignment = header_alignment
        header_style.borders = xlwt.Borders()
        header_style.borders.left = xlwt.Borders.THIN
        header_style.borders.right = xlwt.Borders.THIN
        header_style.borders.top = xlwt.Borders.THIN
        header_style.borders.bottom = xlwt.Borders.THIN

        data_alignment = xlwt.Alignment()
        data_alignment.horz = xlwt.Alignment.HORZ_LEFT
        data_alignment.vert = xlwt.Alignment.VERT_CENTER
        data_alignment.wrap = xlwt.Alignment.WRAP_AT_RIGHT
        data_style = xlwt.XFStyle()
        data_style.font = xlwt.Font()
        data_style.font.height = 220
        data_style.alignment = data_alignment
        data_style.borders = xlwt.Borders()
        data_style.borders.left = xlwt.Borders.THIN
        data_style.borders.right = xlwt.Borders.THIN
        data_style.borders.top = xlwt.Borders.THIN
        data_style.borders.bottom = xlwt.Borders.THIN

        style_cache = {}

        def _xls_font_from_source(src_font):
            font = xlwt.Font()
            font.name = src_font.name
            font.height = int(src_font.height)
            font.bold = bool(src_font.bold)
            font.italic = bool(src_font.italic)
            font.underline = int(getattr(src_font, "underline_type", 0))
            font.colour_index = int(getattr(src_font, "colour_index", 0x7FFF))
            font.escapement = int(getattr(src_font, "escapement", 0))
            font.family = int(getattr(src_font, "family", 0))
            font.charset = int(getattr(src_font, "character_set", 1))
            font.outline = bool(getattr(src_font, "outline", False))
            font.shadow = bool(getattr(src_font, "shadow", False))
            font.struck_out = bool(getattr(src_font, "struck_out", False))
            return font

        def _source_xls_style(row_idx, col_idx, fallback_style):
            try:
                if row_idx >= ws_src.nrows or col_idx >= ws_src.ncols:
                    return fallback_style
                xf_index = ws_src.cell_xf_index(row_idx, col_idx)
                if xf_index in style_cache:
                    return style_cache[xf_index]

                src_book = ws_src.book
                src_xf = src_book.xf_list[xf_index]
                style = xlwt.XFStyle()

                src_font = src_book.font_list[src_xf.font_index]
                style.font = _xls_font_from_source(src_font)

                alignment = xlwt.Alignment()
                alignment.horz = int(src_xf.alignment.hor_align)
                alignment.vert = int(src_xf.alignment.vert_align)
                alignment.wrap = xlwt.Alignment.WRAP_AT_RIGHT
                style.alignment = alignment

                borders = xlwt.Borders()
                src_border = src_xf.border
                borders.left = int(src_border.left_line_style)
                borders.right = int(src_border.right_line_style)
                borders.top = int(src_border.top_line_style)
                borders.bottom = int(src_border.bottom_line_style)
                borders.left_colour = int(src_border.left_colour_index)
                borders.right_colour = int(src_border.right_colour_index)
                borders.top_colour = int(src_border.top_colour_index)
                borders.bottom_colour = int(src_border.bottom_colour_index)
                style.borders = borders

                pattern = xlwt.Pattern()
                src_background = src_xf.background
                pattern.pattern = int(src_background.fill_pattern)
                pattern.pattern_fore_colour = int(src_background.pattern_colour_index)
                pattern.pattern_back_colour = int(src_background.background_colour_index)
                style.pattern = pattern

                try:
                    style.num_format_str = src_book.format_map[src_xf.format_key].format_str
                except Exception:
                    pass

                style_cache[xf_index] = style
                return style
            except Exception:
                return fallback_style

        def _source_xls_rich_text(row_idx, col_idx, value):
            try:
                runlist = ws_src.rich_text_runlist_map.get((row_idx, col_idx))
                if not runlist:
                    return None

                text = str(ws_src.cell_value(row_idx, col_idx))
                if str(value) != text:
                    return None

                rich_parts = []
                runs = sorted((int(offset), int(font_idx)) for offset, font_idx in runlist)
                if not runs:
                    return None

                cursor = 0
                for idx, (offset, font_idx) in enumerate(runs):
                    if offset > cursor:
                        rich_parts.append(text[cursor:offset])
                    next_offset = runs[idx + 1][0] if idx + 1 < len(runs) else len(text)
                    segment = text[offset:next_offset]
                    if segment:
                        rich_parts.append((segment, _xls_font_from_source(ws_src.book.font_list[font_idx])))
                    cursor = next_offset

                if cursor < len(text):
                    rich_parts.append(text[cursor:])
                return rich_parts if rich_parts else None
            except Exception:
                return None

        merges = []
        covered = set()
        try:
            for rlo, rhi, clo, chi in ws_src.merged_cells:
                # 与表头有交集的合并：在表头内按原样合并；若合并延伸到数据区，只保留表头这一段
                if rlo >= n_header or clo >= col_count:
                    continue
                effective_rhi = min(rhi, n_header)
                if effective_rhi <= rlo:
                    continue
                r2 = effective_rhi - 1
                c2 = min(chi - 1, col_count - 1)
                if c2 < clo:
                    continue
                merges.append((rlo, r2, clo, c2))
                for r in range(rlo, effective_rhi):
                    for c in range(clo, chi):
                        if c < col_count and (r != rlo or c != clo):
                            covered.add((r, c))
        except Exception:
            pass

        top_lefts = {(m[0], m[2]) for m in merges}

        def _write_xls_cell(row_idx, col_idx, value, style, source_row_idx=None):
            source_row_idx = row_idx if source_row_idx is None else source_row_idx
            rich_text = _source_xls_rich_text(source_row_idx, col_idx, value)
            if rich_text:
                ws.write_rich_text(row_idx, col_idx, rich_text, style)
            else:
                ws.write(row_idx, col_idx, value, style)

        for r1, r2, c1, c2 in merges:
            row_vals = header_rows[r1] if r1 < len(header_rows) else []
            val = row_vals[c1] if c1 < len(row_vals) else ""
            try:
                rich_text = _source_xls_rich_text(r1, c1, val)
                style = _source_xls_style(r1, c1, header_style)
                if rich_text:
                    ws.write_merge(r1, r2, c1, c2, rich_text, style)
                else:
                    ws.write_merge(r1, r2, c1, c2, val, style)
            except Exception:
                pass

        for r in range(n_header):
            row_vals = header_rows[r] if r < len(header_rows) else []
            for c in range(col_count):
                if (r, c) in covered or (r, c) in top_lefts:
                    continue
                val = row_vals[c] if c < len(row_vals) else ""
                _write_xls_cell(r, c, val, _source_xls_style(r, c, header_style))
            row_obj = ws.row(r)
            height = 450
            try:
                ri = ws_src.rowinfo_map.get(r)
                if ri and ri.height > 0:
                    height = int(ri.height)
            except Exception:
                pass
            row_obj.height = height
            row_obj.height_mismatch = 1

        def _col_width(col_idx):
            ci = ws_src.colinfo_map.get(col_idx)
            if ci is not None and ci.width > 0:
                return max(64, min(ci.width, 65535))
            if ws_src.standardwidth:
                return max(64, min(ws_src.standardwidth, 65535))
            return max(64, min((ws_src.defcolwidth or 8) * 256, 65535))

        def _xls_col_width_chars(col_idx):
            return max(4.0, _col_width(col_idx) / 256.0)

        def _xls_font_size_points(row_idx, col_idx):
            try:
                if row_idx >= ws_src.nrows or col_idx >= ws_src.ncols:
                    return 11
                xf_index = ws_src.cell_xf_index(row_idx, col_idx)
                src_xf = ws_src.book.xf_list[xf_index]
                src_font = ws_src.book.font_list[src_xf.font_index]
                return max(1.0, float(src_font.height) / 20.0)
            except Exception:
                return 11

        def _xls_font_name(row_idx, col_idx):
            try:
                if row_idx >= ws_src.nrows or col_idx >= ws_src.ncols:
                    return None
                xf_index = ws_src.cell_xf_index(row_idx, col_idx)
                src_xf = ws_src.book.xf_list[xf_index]
                src_font = ws_src.book.font_list[src_xf.font_index]
                return src_font.name or None
            except Exception:
                return None

        for i, row_data in enumerate(data_rows):
            rr = n_header + i
            source_rr = data_source_rows[i] if i < len(data_source_rows) else None
            if source_rr is None:
                source_rr = n_header
            for c in range(col_count):
                val = row_data[c] if c < len(row_data) else ""
                _write_xls_cell(rr, c, val, _source_xls_style(source_rr, c, data_style), source_rr)
            row_obj = ws.row(rr)
            estimated_height = self._estimate_row_height(
                row_data,
                col_count,
                _xls_col_width_chars,
                lambda col_idx, row_idx=source_rr: _xls_font_size_points(row_idx, col_idx),
                lambda col_idx, row_idx=source_rr: _xls_font_name(row_idx, col_idx),
            )
            row_obj.height = int(
                estimated_height * 20
            )
            row_obj.height_mismatch = 1

        for col_idx in range(col_count):
            ws.col(col_idx).width = _col_width(col_idx)

        # xlwt 无 page_setup / print_area 属性（勿用 openpyxl 写法）
        # 横向、A4、整表宽度缩放到 1 页宽（行可多页），无页眉页脚，上下顶格
        ws.portrait = 0  # 0=横向
        ws.paper_size_code = 9  # A4
        ws.fit_width_to_pages = 1
        ws.fit_height_to_pages = 0
        ws.header_str = b""
        ws.footer_str = b""
        ws.top_margin = 0
        ws.bottom_margin = 0
        ws.header_margin = 0
        ws.footer_margin = 0

        base_name = os.path.splitext(os.path.basename(b_path))[0]
        date_str = datetime.datetime.now().strftime("%Y%m%d")
        output_filename = f"{base_name}_{date_str}更新.xls"
        output_path = os.path.join(os.path.dirname(b_path), output_filename)
        wb.save(output_path)
        return output_path


def main():
    root = tkinterdnd2.TkinterDnD.Tk()
    
    app = App(root)
    
    root.mainloop()


if __name__ == "__main__":
    main()
